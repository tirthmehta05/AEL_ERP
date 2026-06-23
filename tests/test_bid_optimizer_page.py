"""Tests for pages/bid_optimizer.py.

The page imports streamlit at module load, but conftest.py mocks the entire
streamlit module — so we can import the page in tests and exercise its
pure data-shaping helpers (`_enrich_record`, `_build_excel_blob`) without
launching a real Streamlit runtime.

The hard end-to-end test (Streamlit AppTest with file upload + button click)
lives in Phase F manual testing — running pytest with mocked streamlit will
exercise everything except the actual widget rendering.
"""

from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

import pytest

# conftest.py has already prepended PROJECT_ROOT and mocked streamlit.
from pages import bid_optimizer as bo


# Engine modules were loaded as a side-effect of importing bid_optimizer
# (it prepends slitting_optimizer/ to sys.path). Re-use those here.
from engine import optimizer as eng                # noqa: E402
from engine import incremental as inc              # noqa: E402
from app.repository import auction_repository as ar  # noqa: E402
from app.repository import customer_repository as cr  # noqa: E402
from tools import measurement_report as mr         # noqa: E402

_SLOPT = Path(__file__).resolve().parent.parent / "slitting_optimizer"
# Sanitized FAKE-data fixture — the real customer rates live only in Google
# Sheets (repo is public). Tests assert structure/logic, never specific rates.
_CUSTOMERS = _SLOPT / "data" / "sample_customers_fixture.xlsx"
_MI_PUNE = _SLOPT / "data" / "sample_auction_lists" / "CRNO 27.05.2026 MI PUNE.xlsx"


@pytest.fixture(scope="module")
def orders_and_flags():
    if not _CUSTOMERS.exists():
        pytest.skip(f"customer fixture not found at {_CUSTOMERS}")
    orders, flags = cr.parse_customers(str(_CUSTOMERS))
    return orders, flags


@pytest.fixture(scope="module")
def solved_mi_pune(orders_and_flags):
    """Solve the small MI Pune auction once per test module (~2s) against the
    sanitized fake customer book. Exact allocations/bidability depend on the
    fake rates, so tests below assert STRUCTURE, not specific outcomes."""
    if not _MI_PUNE.exists():
        pytest.skip(f"auction fixture not found at {_MI_PUNE}")
    orders, _ = orders_and_flags
    os.environ.setdefault("SLIT_KNIFE_MAX_WIDE", "19")
    coils, auction_flags = ar.parse_auction(_MI_PUNE)
    from collections import defaultdict
    by_lot = defaultdict(list)
    for c in coils:
        by_lot[c["lot"]].append(c)
    records = [mr._build_lot_record(lot, by_lot[lot], orders, 30)
               for lot in sorted(by_lot)]
    return records, auction_flags


def test_enrich_record_produces_required_ui_fields(solved_mi_pune, orders_and_flags):
    records, _ = solved_mi_pune
    orders, _ = orders_and_flags
    # Any feasible lot exercises the full enrichment path.
    target = next((r for r in records if r["feasible"]), None)
    assert target is not None, "at least one lot must solve feasibly"

    enriched = bo._enrich_record(target, orders)
    # Required UI keys
    for key in ("lot", "weight_kg", "n_coils", "start", "coatings",
                "revenue", "slit_cost", "transport", "tiers",
                "primary_name", "primary_bid", "bidable", "headroom",
                "profit_primary", "margin_net", "margin_gross",
                "transport_pct", "by_cust", "scrap_kg", "scrap_pct",
                "coils", "coil_break"):
        assert key in enriched, f"missing UI field: {key}"
    # Four tiers by default; first is "safe"
    assert len(enriched["tiers"]) == 4
    assert enriched["primary_name"] == "safe"


def test_tier_ordering(solved_mi_pune, orders_and_flags):
    """For any bidable lot: ceiling > aggressive > compete > safe."""
    records, _ = solved_mi_pune
    orders, _ = orders_and_flags
    for r in records:
        if not r["feasible"]:
            continue
        enriched = bo._enrich_record(r, orders)
        bids = [b for _, _, b in enriched["tiers"]]
        assert bids == sorted(bids), f"lot {r['lot']} tiers not ascending: {bids}"


def test_bidable_flag_matches_primary_vs_start(solved_mi_pune, orders_and_flags):
    records, _ = solved_mi_pune
    orders, _ = orders_and_flags
    for r in records:
        if not r["feasible"]:
            continue
        enriched = bo._enrich_record(r, orders)
        expected = enriched["primary_bid"] >= enriched["start"]
        assert enriched["bidable"] == expected, (
            f"lot {r['lot']}: bidable mismatch")


def test_excel_blob_writes_valid_xlsx(solved_mi_pune, orders_and_flags):
    records, auction_flags = solved_mi_pune
    orders, customer_flags = orders_and_flags
    blob = bo._build_excel_blob(records, orders, customer_flags,
                                auction_flags, "MI_PUNE")
    assert isinstance(blob, bytes)
    assert len(blob) > 10_000, "Excel file looks suspiciously small"
    # Verify openpyxl can parse it and finds the expected sheets
    from openpyxl import load_workbook
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(blob)
        wb = load_workbook(f.name, data_only=True)
        os.unlink(f.name)
    expected = {"Summary", "Per-Lot Summary", "Lot P&L by Tier",
                "Lot Detail (Cut Sheets)", "Customer Fulfillment",
                "Customer × Lot", "Bidable Auction Summary",
                "Salvage", "Width Bands", "Flags"}
    assert expected.issubset(set(wb.sheetnames)), (
        f"missing sheets: {expected - set(wb.sheetnames)}")


def test_kapson_dominant_lot_flagged_single_buyer(solved_mi_pune, orders_and_flags):
    """A lot where one customer takes ≥50% should produce a risk-flag string."""
    records, _ = solved_mi_pune
    orders, _ = orders_and_flags
    for r in records:
        if not r["feasible"]:
            continue
        enriched = bo._enrich_record(r, orders)
        flags_html = bo._render_risk_flags(enriched)
        if not enriched["by_cust"]:
            continue
        top_kg = max(enriched["by_cust"].values())
        if top_kg / enriched["weight_kg"] >= 0.5:
            assert "single buyer" in flags_html, (
                f"lot {enriched['lot']} should flag single-buyer risk")


def test_render_helpers_handle_all_lots(solved_mi_pune, orders_and_flags):
    """Render/enrich helpers must not crash on ANY feasible lot, whether it
    has rich customer allocation or none (a SKIP lot)."""
    records, _ = solved_mi_pune
    orders, _ = orders_and_flags
    feasible = [r for r in records if r["feasible"]]
    assert feasible, "fixture should yield at least one feasible lot"
    for r in feasible:
        enriched = bo._enrich_record(r, orders)
        assert isinstance(enriched["bidable"], bool)
        # returns an HTML string (possibly empty) — must not raise
        assert isinstance(bo._render_risk_flags(enriched), str)


def test_quality_haircut_lowers_revenue(solved_mi_pune, orders_and_flags):
    """Applying a per-coil quality < 100% must lower effective revenue and
    therefore every tier bid, vs the same lot at 100% quality."""
    records, _ = solved_mi_pune
    orders, _ = orders_and_flags
    target = next((r for r in records if r["feasible"] and r["coils"]), None)
    assert target is not None
    base = bo._enrich_record(target, orders)
    # Haircut the first coil to 50%.
    cid = target["coils"][0]["id"]
    quality = {cid: {"quality": 0.5, "year": None}}
    cut = bo._enrich_record(target, orders, coil_quality=quality)
    assert cut["revenue"] <= base["revenue"], "quality<100% must not raise revenue"
    # If that coil carried revenue, the safe-tier bid should drop.
    if cut["revenue"] < base["revenue"]:
        assert cut["tiers"][0][2] < base["tiers"][0][2], (
            "lower revenue must lower the safe-tier bid")


def test_transport_override_lowers_bid(solved_mi_pune, orders_and_flags):
    """Raising the source→KAPSON transport rate must not raise any tier bid."""
    records, _ = solved_mi_pune
    orders, _ = orders_and_flags
    target = next((r for r in records if r["feasible"]), None)
    assert target is not None
    base = bo._enrich_record(target, orders,
                             transport_override={"kapson_rs_per_kg": 1.0})
    pricier = bo._enrich_record(target, orders,
                                transport_override={"kapson_rs_per_kg": 5.0})
    assert pricier["tiers"][0][2] <= base["tiers"][0][2], (
        "higher transport must not raise the safe-tier bid")
