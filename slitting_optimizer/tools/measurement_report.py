"""Tier 1 measurement report — runs the validated engine (Increment 1: width
bands + salvage) against sample auction lists and the customer workbook, and
writes a per-auction Excel + a cross-auction summary.

Purpose
-------
Inform — with real data, before adding more model complexity — whether the
gaps we've identified (min transport qty, transport bracket cost, inventory
carryover, origin-aware slit forbid) actually bite, and at what magnitude.
Pure read-only against today's engine; no engine change, no commits, no
mutation of inputs.

Usage
-----
    uv run python -m tools.measurement_report \\
        --customers data/sample_customers_v2.xlsx \\
        --auctions-dir data/sample_auction_lists \\
        --output-dir data/measurement_reports \\
        --time-limit-s 60

`--time-limit-s` default is 60 (fast first look). Bump to 300 for "real"
numbers once the format is signed off. The engine reads other knobs
(`SLIT_BAND_MM`, `SLIT_KNIFE_MAX_WIDE`, `SLIT_SALVAGE`, …) from env as usual.

Output
------
One workbook per auction (7 sheets — see plan in chat 2026-05-22), plus a
`cross_auction_summary.xlsx` on top.
"""

from __future__ import annotations

import argparse
import os
import sys
import time
from collections import defaultdict
from pathlib import Path

# Make `app.*` / `engine.*` importable when run as `python -m tools.…`.
THIS = Path(__file__).resolve()
sys.path.insert(0, str(THIS.parent.parent))

from app.repository import auction_repository as ar  # noqa: E402
from app.repository import customer_repository as cr  # noqa: E402
from app.core.config import settings  # noqa: E402
from engine import incremental as inc  # noqa: E402
from engine import optimizer as eng  # noqa: E402

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402


# ── buckets / thresholds ─────────────────────────────────────────────────────
# All internal quantities are in KG. Bucket ranges are in KG too. (Earlier
# bug: ranges were in tonnes vs kg values → everything > 20 kg mapped to
# ">=20T". Caught by Tirth — now corrected.)

ALLOC_BUCKETS_KG = [(0, 2000), (2000, 5000), (5000, 10000),
                    (10000, 12000), (12000, 15000), (15000, 20000),
                    (20000, 1e12)]
ALLOC_LABELS = ["<2T", "2-5T", "5-10T", "10-12T", "12-15T",
                "15-20T", ">=20T"]
TRANSPORT_BUCKETS_KG = [(0, 5000), (5000, 10000), (10000, 12000),
                         (12000, 20000), (20000, 1e12)]
TRANSPORT_LABELS = ["<=5T", "5-10T", "10-12T", "12-20T", ">=20T"]
WIDTH_BANDS = [(0, 650), (650, 1000), (1000, 1300), (1300, 1e9)]
WIDTH_LABELS = ["<=650mm", "651-1000mm", "1001-1300mm", ">1300mm"]
# 12000 added per user request (12 T = the operational truck-min for this
# project). Both per-lot AND per-(customer × auction) totals checked.
MIN_QTY_THRESHOLDS_KG = [5000, 8000, 10000, 12000, 15000, 20000]

# ── transport cost table (₹/tonne shipped, by truck-size bracket) ────────────
# Provided by Tirth on 2026-05-22. Customer name matched case-insensitive
# substring. Empty list = customer pickup / FOB (we pay nothing).
# (qty_T, rate_₹/T). Customer with multiple entries: smallest bracket ≥ actual
# tonnes is chosen; rate × actual_kg gives the cost. Sub-smallest tonnage
# still uses the smallest available bracket's rate × actual_kg (a literal
# reading of "per tonne"; if these are full-truck flat rates instead, the
# numbers below will need rescaling — please confirm).
TRANSPORT_RATES = {
    "amba enterprises":         [(12, 1000), (19, 1150)],
    "cryotech":                 [(12, 1400)],
    "kalburgi":                 [(12, 1400), (19, 1900)],
    "trimurti stamping":        [(19, 1900)],
    "precision stamping pune":  [(12, 1000), (19, 1150)],
    "silicon cortech":          [(12, 2400), (19, 2400), (30, 2100)],
    "pooja stampings":          [],   # customer pickup / FOB
    "mini":                     [],   # customer pickup / FOB
    "kapson":                   [(999, 1000)],   # flat ₹1/kg (= ₹1000/T)
}


def _transport_cost(customer: str, kg: float) -> tuple[float, str]:
    """Return (cost_₹, bracket_label). Cost is rate × actual_kg (per-tonne).
    Bracket label says which row applied."""
    if kg <= 0:
        return 0.0, "(no shipment)"
    name = customer.lower().strip()
    rates = None
    for key, rs in TRANSPORT_RATES.items():
        if key in name or name in key:
            rates = rs
            break
    if rates is None:
        return 0.0, "(no transport data)"
    if not rates:
        return 0.0, "(customer pickup / FOB)"
    rates_sorted = sorted(rates, key=lambda x: x[0])
    tonnes = kg / 1000
    for qty, rate in rates_sorted:
        if tonnes <= qty:
            return rate * tonnes, f"{qty}T @ ₹{rate}/T"
    qty, rate = rates_sorted[-1]
    return rate * tonnes, f"{qty}T @ ₹{rate}/T (over largest)"


def _bracket_rate_for_auction(customer: str, auction_wide_kg: float) -> float:
    """₹/tonne rate for this customer, given their AUCTION-WIDE total kg.
    Returns 0 for FOB / no-data customers. Used to attribute transport to
    individual lot allocations (the truck is sized by auction-wide total)."""
    if auction_wide_kg <= 0:
        return 0.0
    name = customer.lower().strip()
    rates = None
    for key, rs in TRANSPORT_RATES.items():
        if key in name or name in key:
            rates = rs
            break
    if not rates:
        return 0.0
    rates_sorted = sorted(rates, key=lambda x: x[0])
    tonnes = auction_wide_kg / 1000
    for qty, rate in rates_sorted:
        if tonnes <= qty:
            return rate
    return rates_sorted[-1][1]


def _lot_transport_cost(record, orders, cust_totals_by_name) -> float:
    """Per-lot transport cost: each customer's per-lot allocation × the rate
    determined by their AUCTION-WIDE total (truck size set by total flow)."""
    order_by_id = {o["id"]: o for o in orders}
    total = 0.0
    for oid, kg in record.get("cust_alloc_kg", {}).items():
        cust = order_by_id[oid]["customer"]
        auction_total = cust_totals_by_name.get(cust, 0)
        rate = _bracket_rate_for_auction(cust, auction_total)
        total += rate * kg / 1000  # ₹/T × T
    return total


# Bid-tier helpers live in engine.bid_tiers (single source of truth shared
# with the Streamlit page). Re-exported here under the original private names
# for backward-compat with the rest of this module.
from engine.bid_tiers import (                          # noqa: E402
    bid_for_net_margin as _bid_for_net_margin,
    get_bid_tiers as _get_bid_tiers,
    parse_bid_tiers as _parse_bid_tiers,                # noqa: F401
)


def _bucket(v: float, ranges, labels) -> str:
    for (lo, hi), lab in zip(ranges, labels):
        if lo <= v < hi:
            return lab
    return labels[-1]


# ── styling helpers ──────────────────────────────────────────────────────────

H_FONT = Font(bold=True, color="FFFFFF")
H_FILL = PatternFill("solid", fgColor="305496")
SUB_FONT = Font(bold=True)
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")


def _write_header(ws, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _write_subheader(ws, row: int, label: str, span: int = 1) -> None:
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = SUB_FONT
    cell.fill = SUB_FILL
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=span)


def _autofit(ws, max_w: int = 42) -> None:
    """Best-effort column-width autosizing."""
    widths: dict[int, int] = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(
                widths.get(cell.column, 8),
                min(len(str(cell.value)) + 2, max_w),
            )
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ── per-lot solve & extraction ───────────────────────────────────────────────

def _coil_breakdown(coil, orders, res, s) -> dict:
    """For one coil, return {disposition, cut_pattern, cust_strips, scrap_kg,
    salvage_rev_rs}. cut_pattern is a human-readable slit description;
    cust_strips is {customer_name: strip_kg}."""
    cid = coil["id"]
    wt_kg = coil["weight_g"] / 1000
    width_cdmm = coil["width_cdmm"]

    if s.Value(res["salvage"][cid]):
        delta = res.get("salv_map", {}).get(coil["coating"], 0)
        salv_rate = coil["price_per_kg"] - delta
        return {
            "disposition": "salvage",
            "cut_pattern": f"Salvage @ ₹{salv_rate:.2f}/kg "
                           f"({coil['coating']}, start − ₹{delta:g})",
            "cust_strips": {},
            "scrap_kg": 0.0,
            "salvage_rev_rs": salv_rate * wt_kg,
        }

    if s.Value(res["as_is"][cid]):
        dest = None
        # Path 1: range-match asis order (e.g. KAPSON)
        asis_alloc = res.get("asis_alloc", {})
        for o in orders:
            if o.get("mode") == "asis" \
                    and (cid, o["id"]) in asis_alloc \
                    and s.Value(asis_alloc[cid, o["id"]]) > 0:
                dest = o["customer"]
                break
        # Path 2: exact-width slit match (legacy as-is)
        if not dest:
            for o in orders:
                if (cid, o["id"]) in res["x"] \
                        and s.Value(res["x"][cid, o["id"]]) > 0:
                    dest = o["customer"]
                    break
        return {
            "disposition": "as-is",
            "cut_pattern": f"Whole {width_cdmm / eng.WIDTH_SCALE:g}mm "
                           f"→ {dest or '(none)'}",
            "cust_strips": {dest: wt_kg} if dest else {},
            "scrap_kg": 0.0,
            "salvage_rev_rs": 0.0,
        }

    # slit — group strips by customer for readability (one segment per customer)
    cust_groups: dict[str, list[tuple[int, float]]] = {}  # cust → [(n, mm)]
    cust_strips: dict[str, float] = {}
    for o in orders:
        if (cid, o["id"]) not in res["x"]:
            continue
        n = s.Value(res["x"][cid, o["id"]])
        if n <= 0:
            continue
        w_mm = o["width_cdmm"] / eng.WIDTH_SCALE
        strip_kg = n * (o["width_cdmm"] / width_cdmm) * wt_kg
        cust_groups.setdefault(o["customer"], []).append((n, w_mm))
        cust_strips[o["customer"]] = (
            cust_strips.get(o["customer"], 0) + strip_kg)
    scrap_cdmm = s.Value(res["y"][cid])
    scrap_kg = scrap_cdmm / width_cdmm * wt_kg
    if not cust_groups:
        pattern = "(no strips)"
    else:
        parts = []
        for cust in sorted(cust_groups):
            strips = sorted(cust_groups[cust], key=lambda x: -x[1])
            total_n = sum(n for n, _ in strips)
            total_mm = sum(n * w for n, w in strips)
            inner = " + ".join(f"{n}×{w:g}mm" for n, w in strips)
            parts.append(
                f"{cust} [{total_n} strips, {total_mm:g}mm]: {inner}")
        pattern = "  ·  ".join(parts)
    pattern += (f"  ·  scrap {scrap_cdmm / eng.WIDTH_SCALE:.2f}mm "
                f"({scrap_kg:.1f}kg)")
    return {
        "disposition": "slit",
        "cut_pattern": pattern,
        "cust_strips": cust_strips,
        "scrap_kg": scrap_kg,
        "salvage_rev_rs": 0.0,
    }


def _build_lot_record(lot, coils_in, orders, time_limit_s):
    """Solve one lot vs the FULL customer book (standalone pre-auction view)
    and return everything the report needs in one record."""
    t0 = time.time()
    coils, res, m = inc.price_lot(lot, coils_in, orders)
    elapsed = time.time() - t0
    rec = {
        "lot": lot, "coils": coils, "metrics": m, "res": res,
        "solve_time_s": elapsed,
        "feasible": bool(m.get("feasible")),
        "status": m.get("status"),
    }
    if not rec["feasible"]:
        return rec

    s = res["solver"]

    # Per-coil disposition
    disp = {}
    for c in coils:
        cid = c["id"]
        if s.Value(res["salvage"][cid]):
            disp[cid] = "salvage"
        elif s.Value(res["slit"][cid]):
            disp[cid] = "slit"
        elif s.Value(res["as_is"][cid]):
            disp[cid] = "as-is"
        else:
            disp[cid] = "?"
    rec["disp"] = disp

    # Per-order allocations (kg)
    rec["cust_alloc_kg"] = {oid: g / 1000 for oid, g in m["cust"].items()
                            if g > 0}
    rec["inv_alloc_kg"] = {oid: g / 1000 for oid, g in m["inv"].items()
                           if g > 0}

    # Disposition kg totals
    rec["kg_slit"] = sum(c["weight_g"] / 1000 for c in coils
                         if disp[c["id"]] == "slit")
    rec["kg_asis"] = sum(c["weight_g"] / 1000 for c in coils
                         if disp[c["id"]] == "as-is")
    rec["kg_salvage"] = sum(c["weight_g"] / 1000 for c in coils
                            if disp[c["id"]] == "salvage")
    rec["kg_cust"] = sum(rec["cust_alloc_kg"].values())
    rec["kg_inv"] = sum(rec["inv_alloc_kg"].values())
    rec["kg_scrap"] = m["scrap_kg"]

    # Slit cost (needed for net-margin-at-max-bid calculation) — recompute
    # since metrics() doesn't return it. Per-coil rate via the width band.
    rec["slit_cost_rs"] = sum(eng.slit_cost_for(c) * c["weight_g"] / 1000
                              for c in coils if disp[c["id"]] == "slit")

    # Per-coil cut-sheet breakdown (cut pattern, customer strip kg, scrap)
    rec["coil_break"] = {c["id"]: _coil_breakdown(c, orders, res, s)
                         for c in coils}

    # Salvage detail
    salv_map = res.get("salv_map", {}) or {}
    salvaged = []
    for c in coils:
        if disp[c["id"]] != "salvage":
            continue
        delta = salv_map.get(c["coating"], 0)
        salv_price = c["price_per_kg"] - delta
        wt_kg = c["weight_g"] / 1000
        salv_rev = salv_price * wt_kg
        scrap_rev = eng.SCRAP_RATE * wt_kg
        salvaged.append({
            "batch": c["batch"], "coating": c["coating"],
            "width_mm": c["width_cdmm"] / eng.WIDTH_SCALE,
            "weight_kg": wt_kg,
            "lot_start_per_kg": c["price_per_kg"],
            "delta": delta, "salv_price_per_kg": salv_price,
            "salv_rev_rs": salv_rev,
            "scrap_counterfactual_rs": scrap_rev,
            "savings_rs": salv_rev - scrap_rev,
        })
    rec["salvaged"] = salvaged

    # Coils by width band
    by_band: dict[str, dict] = {lab: {"count": 0, "kg": 0.0,
                                       "slit_count": 0, "slit_kg": 0.0,
                                       "asis_count": 0, "asis_kg": 0.0,
                                       "salv_count": 0, "salv_kg": 0.0}
                                 for lab in WIDTH_LABELS}
    for c in coils:
        w_mm = c["width_cdmm"] / eng.WIDTH_SCALE
        wt_kg = c["weight_g"] / 1000
        band = _bucket(w_mm, WIDTH_BANDS, WIDTH_LABELS)
        d = disp[c["id"]]
        by_band[band]["count"] += 1
        by_band[band]["kg"] += wt_kg
        if d == "slit":
            by_band[band]["slit_count"] += 1
            by_band[band]["slit_kg"] += wt_kg
        elif d == "as-is":
            by_band[band]["asis_count"] += 1
            by_band[band]["asis_kg"] += wt_kg
        elif d == "salvage":
            by_band[band]["salv_count"] += 1
            by_band[band]["salv_kg"] += wt_kg
    rec["by_band"] = by_band

    return rec


# ── aggregations across lots ─────────────────────────────────────────────────

def _customer_lot_alloc(lot_records, orders):
    """List[{customer, lot, kg, revenue_at_rate, bucket}] — per-(customer,lot)
    rolled up across that customer's multiple orders if any. Used for the
    MIN_QTY histogram + simulation."""
    order_by_id = {o["id"]: o for o in orders}
    by_cl: dict[tuple[str, str], dict] = defaultdict(
        lambda: {"kg": 0.0, "rev": 0.0})
    for r in lot_records:
        if not r["feasible"]:
            continue
        lot = r["lot"]
        for oid, kg in r["cust_alloc_kg"].items():
            o = order_by_id[oid]
            key = (o["customer"], lot)
            by_cl[key]["kg"] += kg
            by_cl[key]["rev"] += kg * o["rate_per_kg"]
    rows = []
    for (cust, lot), v in by_cl.items():
        rows.append({
            "customer": cust, "lot": lot,
            "kg": v["kg"],
            "revenue_at_rate": v["rev"],
            "bucket": _bucket(v["kg"], ALLOC_BUCKETS_KG, ALLOC_LABELS),
        })
    rows.sort(key=lambda r: (r["customer"], r["lot"]))
    return rows


def _customer_totals(cust_lot_alloc, orders):
    """Per-customer auction-wide totals + transport bracket + monthly demand
    for context."""
    demand = defaultdict(float)
    for o in orders:
        demand[o["customer"]] += o["monthly_g"] / 1000

    totals = defaultdict(float)
    rev = defaultdict(float)
    for r in cust_lot_alloc:
        totals[r["customer"]] += r["kg"]
        rev[r["customer"]] += r["revenue_at_rate"]

    rows = []
    customers = set(totals) | set(demand)
    for c in sorted(customers):
        kg = totals.get(c, 0)
        rows.append({
            "customer": c,
            "total_kg": kg,
            "monthly_demand_kg": demand.get(c, 0),
            "bracket": (_bucket(kg, TRANSPORT_BUCKETS_KG, TRANSPORT_LABELS)
                        if kg > 0 else "(none)"),
            "revenue_at_rate": rev.get(c, 0),
        })
    return rows


def _min_qty_sim(rows, kg_key="kg"):
    """For each MIN_QTY threshold (kg), how many entries fall below it.
    Generic — works on per-(customer,lot) rows OR per-customer auction-wide
    rows. `rows` items must have `kg_key` + 'revenue_at_rate' fields."""
    sims = []
    rows = [r for r in rows if r.get(kg_key, 0) > 0]
    total_count = len(rows)
    total_kg = sum(r[kg_key] for r in rows)
    total_rev = sum(r["revenue_at_rate"] for r in rows)
    for thr in MIN_QTY_THRESHOLDS_KG:
        below = [r for r in rows if r[kg_key] < thr]
        sims.append({
            "threshold_kg": thr,
            "below_count": len(below),
            "below_pct": (100 * len(below) / total_count
                          if total_count else 0),
            "below_total_kg": sum(r[kg_key] for r in below),
            "below_pct_of_total_kg": (100 * sum(r[kg_key] for r in below)
                                       / total_kg if total_kg else 0),
            "foregone_rev_rs": sum(r["revenue_at_rate"] for r in below),
            "foregone_pct_of_total_rev": (100 * sum(r["revenue_at_rate"]
                                                     for r in below)
                                           / total_rev if total_rev else 0),
        })
    return sims


def _flags(lot_records, coils_all, orders, auction_flags, customer_flags):
    """Data-hygiene flags: salvage-only coatings in customer file, orders
    with no width-match in the auction, parse anomalies."""
    out = []
    out.extend(f"[parse] {f}" for f in auction_flags)
    out.extend(f"[parse] {f}" for f in customer_flags)

    # Salvage-only coatings appearing in customer orders
    salv_coats = set(settings.salvage_map.keys())
    if salv_coats:
        for o in orders:
            overlap = sorted(o["coatings"] & salv_coats)
            if overlap:
                out.append(
                    f"[salvage-hygiene] Customer '{o['customer']}' (oid "
                    f"{o['id']}, w{o['width_cdmm'] / eng.WIDTH_SCALE:g}mm) "
                    f"lists salvage-only coating(s) {overlap} — remove or "
                    f"bid is inflated.")

    # Coils with zero rate/width/weight (parse anomalies)
    for c in coils_all:
        if c["weight_g"] == 0 or c["width_cdmm"] == 0:
            out.append(f"[coil] lot {c['lot']} batch {c['batch']} — "
                       f"zero width or weight")
        if c["price_per_kg"] == 0:
            out.append(f"[coil] lot {c['lot']} batch {c['batch']} — "
                       f"zero lot start price")

    # Orders with no matching coil anywhere in the auction
    for o in orders:
        any_match = False
        for c in coils_all:
            if (c["grade"] in o["grades"]
                    and c["coating"] in o["coatings"]
                    and c["thickness"] in o["thicknesses"]
                    and c["width_cdmm"] >= o["width_cdmm"]):
                any_match = True
                break
        if not any_match:
            out.append(
                f"[demand-mismatch] Order oid {o['id']} ('{o['customer']}', "
                f"w{o['width_cdmm'] / eng.WIDTH_SCALE:g}mm, "
                f"grades {sorted(o['grades'])}, "
                f"coatings {sorted(o['coatings'])}) — no compatible coil in "
                f"this auction (effectively dead demand here).")
    return out


# ── per-auction workbook writers ─────────────────────────────────────────────

def _write_summary(ws, auction_name, data):
    lots = data["lot_records"]
    feas = [r for r in lots if r["feasible"]]
    total_kg = sum(c["weight_g"] for c in data["coils"]) / 1000
    total_lots = len(lots)

    # Bid / profit headline — primary tier (transport-aware) for the avg bid.
    tiers = _get_bid_tiers()
    primary_name, primary_margin = tiers[0]
    cust_totals_by_name_for_avg = {r["customer"]: r["total_kg"]
                                   for r in data["cust_totals"]}
    profit_total = sum(r["metrics"]["profit"] for r in feas)
    bid_weighted_kg = sum(
        _bid_for_net_margin(r["metrics"]["total_rev"],
                            r.get("slit_cost_rs", 0),
                            _lot_transport_cost(r, data["orders"],
                                                cust_totals_by_name_for_avg),
                            r["metrics"]["total_wt"], primary_margin)
        * r["metrics"]["total_wt"]
        for r in feas
    )
    bid_weight = sum(r["metrics"]["total_wt"] for r in feas)
    avg_bid_per_kg = bid_weighted_kg / bid_weight if bid_weight else 0
    total_rev = sum(r["metrics"]["total_rev"] for r in feas)
    avg_margin_pct = (100 * profit_total / total_rev) if total_rev else 0

    # Dispositions (kg)
    kg_slit = sum(r["kg_slit"] for r in feas)
    kg_asis = sum(r["kg_asis"] for r in feas)
    kg_salvage = sum(r["kg_salvage"] for r in feas)
    kg_inv = sum(r["kg_inv"] for r in feas)
    kg_scrap = sum(r["kg_scrap"] for r in feas)

    # Salvage savings
    salv_savings = sum(s["savings_rs"]
                       for r in feas for s in r["salvaged"])
    salv_count = sum(len(r["salvaged"]) for r in feas)
    kg_cust = sum(r["kg_cust"] for r in feas)

    # Solve stats
    total_solve_s = sum(r["solve_time_s"] for r in lots)
    feas_optimal = sum(1 for r in feas if r["metrics"].get("optimal"))
    avg_gap = (sum(r["metrics"].get("gap_pct") or 0 for r in feas
                   if r["metrics"].get("gap_pct") is not None)
               / max(1, sum(1 for r in feas
                            if r["metrics"].get("gap_pct") is not None)))

    ws.title = "Summary"
    ws.cell(row=1, column=1, value=f"Measurement Report — {auction_name}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Customers: {data['n_customers']}    "
                                   f"Orders: {len(data['orders'])}    "
                                   f"Coils: {len(data['coils'])}    "
                                   f"Lots: {total_lots}    "
                                   f"Total weight: {total_kg:,.0f} kg")

    rows = [
        ("─── Bidding ──", ""),
        ("Total profit @ start (₹)", f"{profit_total:,.0f}"),
        (f"Weighted avg {primary_name} bid (₹/kg, post-transport)",
         f"{avg_bid_per_kg:,.2f}"),
        ("Total revenue (₹)", f"{total_rev:,.0f}"),
        ("Avg margin (%)", f"{avg_margin_pct:.2f}"),
        ("", ""),
        ("─── Dispositions (kg) ──", ""),
        ("kg slit (full coil weight)", f"{kg_slit:,.0f}"),
        ("kg as-is", f"{kg_asis:,.0f}"),
        ("kg salvage", f"{kg_salvage:,.0f}"),
        ("kg to customer (strips)", f"{kg_cust:,.0f}"),
        ("kg to inventory", f"{kg_inv:,.0f}"),
        ("kg scrap", f"{kg_scrap:,.0f}"),
        ("scrap % of total weight", f"{100 * kg_scrap / total_kg:.2f}" if total_kg else "0"),
        ("", ""),
        ("─── Salvage ──", ""),
        ("Coils salvaged", str(salv_count)),
        ("Salvage savings vs scrap (₹)", f"{salv_savings:,.0f}"),
        ("", ""),
        ("─── Solve stats ──", ""),
        ("Lots solved (feasible)", f"{len(feas)} / {total_lots}"),
        ("Lots proven optimal", str(feas_optimal)),
        ("Avg gap % (non-optimal lots)", f"{avg_gap:.2f}"),
        ("Total solve time (s)", f"{total_solve_s:,.1f}"),
        ("", ""),
        ("─── Data-hygiene flags ──", ""),
        ("Flag count", str(len(data["flags"]))),
        ("(see Flags sheet)", ""),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        if k.startswith("───"):
            ws.cell(row=i, column=1).font = SUB_FONT
            ws.cell(row=i, column=1).fill = SUB_FILL
    _autofit(ws)


def _write_per_lot_summary(ws, data):
    """One row per lot. Configurable bid tiers via SLIT_BID_TIERS env;
    the FIRST tier ("primary") drives the bidable flag and the profit/margin
    display columns. All bids are POST-transport (transport per-lot via
    customer auction-wide bracket)."""
    cust_totals_by_name = {r["customer"]: r["total_kg"]
                           for r in data["cust_totals"]}
    orders = data["orders"]
    tiers = _get_bid_tiers()
    primary_name, primary_margin = tiers[0]

    bid_headers = [f"bid_{n}_{int(m*100)}%net_₹/kg" if m > 0
                   else f"bid_{n}_0%_₹/kg"
                   for n, m in tiers]
    static_headers = ["lot", "bidable", "coils", "weight_kg", "start_₹/kg"]
    tail_headers = [
        f"headroom_{primary_name}_₹/kg",
        "kg_slit", "kg_asis", "kg_salvage", "kg_cust",
        "kg_inv", "kg_scrap", "scrap_%",
        "revenue_₹", "slit_cost_₹", "transport_₹", "transport_%_of_rev",
        "profit_@_start_₹", "margin_@_start_%",
        f"profit_@_{primary_name}_₹",
        f"gross_margin_@_{primary_name}_%",
        f"net_margin_@_{primary_name}_%",
        "salvage_count", "salvage_savings_₹",
        "gap_%", "status", "solve_s",
    ]
    headers = static_headers + bid_headers + tail_headers
    _write_header(ws, 1, headers)
    status_col = len(headers) - 1
    solve_col = len(headers)
    for i, r in enumerate(data["lot_records"], start=2):
        if not r["feasible"]:
            ws.cell(row=i, column=1, value=r["lot"])
            ws.cell(row=i, column=status_col, value=f"ERROR/{r['status']}")
            ws.cell(row=i, column=solve_col, value=round(r["solve_time_s"], 1))
            continue
        m = r["metrics"]
        coils = r["coils"]
        wt_kg = m["total_wt"]                       # kg
        start = coils[0]["price_per_kg"] if coils else 0
        revenue = m["total_rev"]
        slit_cost = r.get("slit_cost_rs", 0.0)
        transport = _lot_transport_cost(r, orders, cust_totals_by_name)
        bid_vals = [_bid_for_net_margin(revenue, slit_cost, transport,
                                        wt_kg, mn)
                    for _, mn in tiers]
        bid_primary = bid_vals[0]
        bidable = bid_primary >= start
        profit_at_start = m["profit"] - transport
        margin_at_start = (100 * profit_at_start / revenue) if revenue else 0
        profit_at_primary = (revenue - bid_primary * wt_kg
                             - slit_cost - transport)
        net_margin_primary = ((100 * profit_at_primary / revenue)
                              if revenue else 0)
        gross_margin_primary = net_margin_primary + (
            (100 * transport / revenue) if revenue else 0)
        transport_pct_rev = ((100 * transport / revenue) if revenue else 0)
        salv_savings = sum(s["savings_rs"] for s in r["salvaged"])
        scrap_pct = (100 * r["kg_scrap"] / wt_kg) if wt_kg else 0
        row = (
            [r["lot"], "YES" if bidable else "NO", len(coils),
             round(wt_kg, 1), round(start, 2)]
            + [round(v, 2) for v in bid_vals]
            + [round(bid_primary - start, 2),
               round(r["kg_slit"], 1), round(r["kg_asis"], 1),
               round(r["kg_salvage"], 1), round(r["kg_cust"], 1),
               round(r["kg_inv"], 1), round(r["kg_scrap"], 1),
               round(scrap_pct, 2),
               round(revenue, 0), round(slit_cost, 0), round(transport, 0),
               round(transport_pct_rev, 2),
               round(profit_at_start, 0), round(margin_at_start, 2),
               round(profit_at_primary, 0),
               round(gross_margin_primary, 2),
               round(net_margin_primary, 2),
               len(r["salvaged"]), round(salv_savings, 0),
               (round(m["gap_pct"], 2) if m.get("gap_pct") is not None
                else None),
               m.get("status"), round(r["solve_time_s"], 1)]
        )
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)
        bid_cell = ws.cell(row=i, column=2)
        bid_cell.fill = GOOD_FILL if bidable else BAD_FILL
    _autofit(ws)


def _write_lot_detail(ws, data):
    """One row per (lot, coil) — disposition, cut pattern, customer
    destinations with kg, scrap kg."""
    ws.cell(row=1, column=1, value="Per-coil cut sheet. Each row is one coil with its disposition decision and the slit/destination pattern.").font = Font(italic=True, color="888888")
    headers = ["lot", "batch", "width_mm", "weight_kg", "thickness", "grade",
               "coating", "disposition", "cut_pattern",
               "customer_destinations_kg", "scrap_kg", "salvage_rev_₹"]
    _write_header(ws, 3, headers)
    row_idx = 4
    for r in data["lot_records"]:
        if not r["feasible"]:
            continue
        cb = r.get("coil_break", {})
        for c in r["coils"]:
            bk = cb.get(c["id"], {})
            cust_str = "; ".join(f"{cust}({kg:.1f}kg)"
                                  for cust, kg
                                  in sorted(bk.get("cust_strips", {}).items()))
            row = [
                r["lot"], c["batch"], c["width_cdmm"] / eng.WIDTH_SCALE,
                round(c["weight_g"] / 1000, 1), c["thickness"],
                c["grade"], c["coating"],
                bk.get("disposition", "?"),
                bk.get("cut_pattern", ""),
                cust_str,
                round(bk.get("scrap_kg", 0), 1),
                round(bk.get("salvage_rev_rs", 0), 0)
                  if bk.get("disposition") == "salvage" else None,
            ]
            for j, v in enumerate(row, start=1):
                ws.cell(row=row_idx, column=j, value=v)
            # Color disposition column
            dc = ws.cell(row=row_idx, column=8)
            if bk.get("disposition") == "salvage":
                dc.fill = WARN_FILL
            elif bk.get("disposition") == "as-is":
                dc.fill = SUB_FILL
            row_idx += 1
    _autofit(ws)


def _write_lot_pnl_detail(ws, data):
    """One block per lot — full P&L table at each configured bid tier.
    Shows: bid ₹/kg, lot cost ₹ (=bid×wt), revenue ₹, slit cost ₹, transport ₹,
    gross profit ₹ (pre-transport), gross margin %, net profit ₹ (post-transport),
    net margin %. The bidder sees the per-tier money flow at a glance."""
    cust_totals_by_name = {r["customer"]: r["total_kg"]
                           for r in data["cust_totals"]}
    orders = data["orders"]
    tiers = _get_bid_tiers()
    primary_name, _ = tiers[0]

    ws.cell(row=1, column=1, value="Per-lot P&L breakdown at each bid tier — bid (₹/kg), lot cost, revenue, slit cost, transport, gross & net profit/margin. Bidable = primary tier bid ≥ start price.").font = Font(italic=True, color="888888")

    row = 3
    for r in data["lot_records"]:
        if not r["feasible"]:
            ws.cell(row=row, column=1, value=f"Lot {r['lot']} — INFEASIBLE ({r['status']})").font = SUB_FONT
            row += 2
            continue
        m = r["metrics"]
        coils = r["coils"]
        wt_kg = m["total_wt"]
        start = coils[0]["price_per_kg"] if coils else 0
        revenue = m["total_rev"]
        slit_cost = r.get("slit_cost_rs", 0.0)
        transport = _lot_transport_cost(r, orders, cust_totals_by_name)

        # Lot header row
        bid_primary = _bid_for_net_margin(revenue, slit_cost, transport,
                                          wt_kg, tiers[0][1])
        bidable = bid_primary >= start
        coatings = sorted({c["coating"] for c in coils})
        header_txt = (f"Lot {r['lot']}  ·  {wt_kg/1000:.1f} T  ·  "
                      f"{len(coils)} coils  ·  start ₹{start:g}/kg  ·  "
                      f"coatings: {','.join(coatings)}  ·  "
                      f"{'✅ BID' if bidable else '❌ SKIP'}")
        hc = ws.cell(row=row, column=1, value=header_txt)
        hc.font = Font(bold=True)
        hc.fill = GOOD_FILL if bidable else BAD_FILL
        row += 1

        # Per-tier P&L table
        _write_header(ws, row, ["tier", "target_net_margin_%",
                                "bid_₹/kg", "lot_cost_₹", "revenue_₹",
                                "slit_cost_₹", "transport_₹",
                                "gross_profit_₹", "gross_margin_%",
                                "net_profit_₹", "net_margin_%"])
        row += 1
        for name, mn in tiers:
            bid = _bid_for_net_margin(revenue, slit_cost, transport, wt_kg, mn)
            lot_cost = bid * wt_kg
            gross_profit = revenue - lot_cost - slit_cost
            gross_margin = (100 * gross_profit / revenue) if revenue else 0
            net_profit = gross_profit - transport
            net_margin = (100 * net_profit / revenue) if revenue else 0
            vals = [name, round(mn * 100, 2),
                    round(bid, 2), round(lot_cost, 0),
                    round(revenue, 0), round(slit_cost, 0),
                    round(transport, 0),
                    round(gross_profit, 0), round(gross_margin, 2),
                    round(net_profit, 0), round(net_margin, 2)]
            for j, v in enumerate(vals, start=1):
                ws.cell(row=row, column=j, value=v)
            if name == primary_name:
                for j in range(1, 12):
                    ws.cell(row=row, column=j).fill = SUB_FILL
            row += 1
        row += 1   # blank separator
    _autofit(ws)


def _write_customer_fulfillment(ws, data):
    """Per-customer drill-down. For each customer: total kg this auction,
    then per (lot, coil) — coil batch / width / coating, the strip pattern
    they receive, and their kg from that coil."""
    orders = data["orders"]
    order_by_id = {o["id"]: o for o in orders}

    # customer → list of (lot, coil_obj, strip_pattern_for_this_cust, cust_kg)
    cust_rows: dict[str, list[tuple]] = defaultdict(list)
    cust_totals: dict[str, float] = defaultdict(float)
    for r in data["lot_records"]:
        if not r["feasible"]:
            continue
        cb = r.get("coil_break", {})
        for c in r["coils"]:
            bk = cb.get(c["id"], {})
            cust_strips = bk.get("cust_strips", {})
            if not cust_strips:
                continue
            # Per-customer slice of this coil's cut pattern
            cust_strip_kg = cust_strips
            for cust, kg in cust_strip_kg.items():
                cust_rows[cust].append((r["lot"], c, bk, kg))
                cust_totals[cust] += kg

    ws.cell(row=1, column=1, value="Customer fulfillment drill-down — per customer, every (lot, coil) they receive material from with widths and kg.").font = Font(italic=True, color="888888")

    row = 3
    for cust in sorted(cust_rows):
        rows_for_cust = cust_rows[cust]
        total_kg = cust_totals[cust]
        n_lots = len({lot for lot, _, _, _ in rows_for_cust})
        n_coils = len(rows_for_cust)
        # Section header
        hc = ws.cell(
            row=row, column=1,
            value=f"{cust}  ·  total {total_kg:,.0f} kg  ·  "
                  f"{n_coils} coils across {n_lots} lots")
        hc.font = Font(bold=True)
        hc.fill = SUB_FILL
        row += 1
        # Table
        _write_header(ws, row, ["lot", "batch", "width_mm", "weight_kg",
                                "grade", "coating", "disposition",
                                "this_customer_pattern", "kg_to_this_customer"])
        row += 1
        for lot, c, bk, kg in sorted(rows_for_cust, key=lambda x: x[0]):
            # Reconstruct just THIS customer's portion of the cut pattern
            pat = bk.get("cut_pattern", "")
            if bk.get("disposition") == "as-is":
                this_pat = f"whole coil → {cust}"
            elif bk.get("disposition") == "slit":
                # Extract this customer's segment from the grouped pattern
                seg = [p for p in pat.split("  ·  ")
                       if p.startswith(cust + " ")]
                this_pat = seg[0] if seg else "(see lot cut sheet)"
            else:
                this_pat = pat
            vals = [lot, c["batch"], c["width_cdmm"] / eng.WIDTH_SCALE,
                    round(c["weight_g"] / 1000, 1),
                    c["grade"], c["coating"],
                    bk.get("disposition", "?"),
                    this_pat, round(kg, 1)]
            for j, v in enumerate(vals, start=1):
                ws.cell(row=row, column=j, value=v)
            row += 1
        row += 1   # blank separator
    _autofit(ws)


def _write_bidable_summary(ws, data):
    """Aggregate across bidable lots only — grand totals (per tier) +
    per-customer + per-lot breakdown. `bidable` = primary tier bid ≥ start
    price. Configurable tiers via SLIT_BID_TIERS env."""
    cust_totals_by_name = {r["customer"]: r["total_kg"]
                           for r in data["cust_totals"]}
    orders = data["orders"]
    tiers = _get_bid_tiers()
    primary_name, primary_margin = tiers[0]

    def _bid_per_lot(r, margin_net):
        wt = r["metrics"]["total_wt"]
        rev = r["metrics"]["total_rev"]
        sc = r.get("slit_cost_rs", 0)
        tr = _lot_transport_cost(r, orders, cust_totals_by_name)
        return _bid_for_net_margin(rev, sc, tr, wt, margin_net)

    # Filter to bidable lots (primary tier ≥ start)
    bidable = [r for r in data["lot_records"]
               if r["feasible"]
               and r["coils"]
               and _bid_per_lot(r, primary_margin)
                   >= r["coils"][0]["price_per_kg"]]

    if not bidable:
        ws.cell(row=1, column=1,
                value="No bidable lots in this auction.").font = SUB_FONT
        _autofit(ws)
        return

    order_by_id = {o["id"]: o for o in orders}
    cust_kg = defaultdict(float)
    cust_gross = defaultdict(float)
    for r in bidable:
        for oid, kg in r["cust_alloc_kg"].items():
            o = order_by_id[oid]
            cust_kg[o["customer"]] += kg
            cust_gross[o["customer"]] += kg * o["rate_per_kg"]
    cust_transport = {}
    for cust, kg in cust_kg.items():
        cost, _ = _transport_cost(cust, kg)
        cust_transport[cust] = cost

    # Grand totals
    total_kg = sum(r["metrics"]["total_wt"] for r in bidable)
    total_rev = sum(r["metrics"]["total_rev"] for r in bidable)
    total_slit_cost = sum(r.get("slit_cost_rs", 0) for r in bidable)
    total_transport = sum(_lot_transport_cost(r, orders, cust_totals_by_name)
                          for r in bidable)
    total_scrap = sum(r["kg_scrap"] for r in bidable)
    total_salvage_kg = sum(r["kg_salvage"] for r in bidable)
    total_bid_cost_start = sum((r["coils"][0]["price_per_kg"]
                                * r["metrics"]["total_wt"]) for r in bidable)
    profit_at_start = (total_rev - total_bid_cost_start
                       - total_slit_cost - total_transport)
    margin_at_start = (100 * profit_at_start / total_rev) if total_rev else 0

    ws.cell(row=1, column=1, value=f"Bidable lots only ({len(bidable)} of {len(data['lot_records'])} lots) — totals + per-customer + per-lot breakdown.").font = SUB_FONT
    ws.cell(row=2, column=1, value=f"Bidable = {primary_name} bid ≥ lot start price. Transport per lot uses customer's auction-wide bracket.").font = Font(italic=True, color="888888")

    grand_row = 4
    _write_subheader(ws, grand_row, "GRAND TOTALS (bidable lots)", span=2)
    rows = [
        ("Lots (bidable / total)", f"{len(bidable)} / {len(data['lot_records'])}"),
        ("Total weight (kg)", f"{total_kg:,.0f}"),
        ("Total revenue (₹)", f"{total_rev:,.0f}"),
        ("Total bid cost @ start (₹)", f"{total_bid_cost_start:,.0f}"),
        ("Total slit cost (₹)", f"{total_slit_cost:,.0f}"),
        ("Total transport (₹)", f"{total_transport:,.0f}"),
        ("Total scrap (kg)", f"{total_scrap:,.0f}"),
        ("Total salvage (kg)", f"{total_salvage_kg:,.0f}"),
        ("──", ""),
        ("Profit @ start (₹)", f"{profit_at_start:,.0f}"),
        ("Margin @ start (%)", f"{margin_at_start:.2f}"),
    ]
    for name, mn in tiers:
        bid_cost = sum(_bid_per_lot(r, mn) * r["metrics"]["total_wt"]
                       for r in bidable)
        profit = total_rev - bid_cost - total_slit_cost - total_transport
        margin = (100 * profit / total_rev) if total_rev else 0
        rows.append((f"Profit @ {name} bid ({int(mn*100)}% net) (₹)",
                     f"{profit:,.0f}"))
        rows.append((f"Net margin @ {name} (%)", f"{margin:.2f}"))
    for i, (k, v) in enumerate(rows, start=grand_row + 1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)

    cust_start = grand_row + len(rows) + 3
    _write_subheader(ws, cust_start, "Per-customer aggregate (bidable lots)",
                     span=6)
    _write_header(ws, cust_start + 1, ["customer", "total_kg", "gross_rev_₹",
                                       "transport_₹", "net_rev_₹",
                                       "transport_%_of_gross"])
    for i, cust in enumerate(sorted(cust_kg), start=cust_start + 2):
        kg = cust_kg[cust]
        gross = cust_gross[cust]
        transport = cust_transport[cust]
        net = gross - transport
        pct = (100 * transport / gross) if gross else 0
        ws.cell(row=i, column=1, value=cust)
        ws.cell(row=i, column=2, value=round(kg, 1))
        ws.cell(row=i, column=3, value=round(gross, 0))
        ws.cell(row=i, column=4, value=round(transport, 0))
        ws.cell(row=i, column=5, value=round(net, 0))
        ws.cell(row=i, column=6, value=round(pct, 2))

    # Per-lot detail: one column per tier
    lot_start = cust_start + 2 + len(cust_kg) + 2
    _write_subheader(ws, lot_start, "Per-lot detail (bidable lots only)",
                     span=5 + len(tiers))
    bid_cols = [f"bid_{n}_₹/kg" for n, _ in tiers]
    _write_header(ws, lot_start + 1, ["lot", "weight_kg", "start_₹/kg"]
                  + bid_cols
                  + ["revenue_₹", "transport_₹",
                     f"profit_@_{primary_name}_₹",
                     f"net_margin_@_{primary_name}_%"])
    for i, r in enumerate(bidable, start=lot_start + 2):
        m = r["metrics"]
        wt_kg = m["total_wt"]
        start = r["coils"][0]["price_per_kg"]
        revenue = m["total_rev"]
        slit_cost = r.get("slit_cost_rs", 0)
        transport = _lot_transport_cost(r, orders, cust_totals_by_name)
        bid_vals = [_bid_for_net_margin(revenue, slit_cost, transport,
                                        wt_kg, mn)
                    for _, mn in tiers]
        bid_primary = bid_vals[0]
        pm = revenue - bid_primary * wt_kg - slit_cost - transport
        mm = (100 * pm / revenue) if revenue else 0
        col = 1
        ws.cell(row=i, column=col, value=r["lot"]); col += 1
        ws.cell(row=i, column=col, value=round(wt_kg, 1)); col += 1
        ws.cell(row=i, column=col, value=round(start, 2)); col += 1
        for v in bid_vals:
            ws.cell(row=i, column=col, value=round(v, 2)); col += 1
        ws.cell(row=i, column=col, value=round(revenue, 0)); col += 1
        ws.cell(row=i, column=col, value=round(transport, 0)); col += 1
        ws.cell(row=i, column=col, value=round(pm, 0)); col += 1
        ws.cell(row=i, column=col, value=round(mm, 2))

    _autofit(ws)


def _write_cust_lot_alloc(ws, data):
    alloc = data["cust_lot_alloc"]

    # Note row
    ws.cell(row=1, column=1, value="Standalone per-lot allocations (no inter-lot demand consistency). Each row = (customer, lot) total kg.").font = Font(italic=True, color="888888")

    # Detail block
    _write_header(ws, 3, ["customer", "lot", "kg", "revenue_at_rate_₹", "bucket"])
    for i, r in enumerate(alloc, start=4):
        ws.cell(row=i, column=1, value=r["customer"])
        ws.cell(row=i, column=2, value=r["lot"])
        ws.cell(row=i, column=3, value=round(r["kg"], 1))
        ws.cell(row=i, column=4, value=round(r["revenue_at_rate"], 0))
        ws.cell(row=i, column=5, value=r["bucket"])

    # Histogram block
    hist_row = len(alloc) + 6
    _write_subheader(ws, hist_row, "Bucket distribution", span=5)
    _write_header(ws, hist_row + 1, ["bucket", "count", "kg_total", "rev_total_₹"])
    bucket_data = defaultdict(lambda: {"count": 0, "kg": 0.0, "rev": 0.0})
    for r in alloc:
        b = r["bucket"]
        bucket_data[b]["count"] += 1
        bucket_data[b]["kg"] += r["kg"]
        bucket_data[b]["rev"] += r["revenue_at_rate"]
    for j, label in enumerate(ALLOC_LABELS, start=hist_row + 2):
        d = bucket_data.get(label, {"count": 0, "kg": 0, "rev": 0})
        ws.cell(row=j, column=1, value=label)
        ws.cell(row=j, column=2, value=d["count"])
        ws.cell(row=j, column=3, value=round(d["kg"], 1))
        ws.cell(row=j, column=4, value=round(d["rev"], 0))

    # MIN_QTY simulation block
    sim_row = hist_row + 2 + len(ALLOC_LABELS) + 2
    _write_subheader(ws, sim_row, "MIN_QTY simulation — allocations BELOW threshold", span=7)
    _write_header(ws, sim_row + 1, ["MIN_QTY_kg", "below_count", "below_%",
                                    "below_total_kg", "below_%_of_total_kg",
                                    "foregone_rev_₹", "foregone_%_of_rev"])
    for j, s in enumerate(data["min_qty_sim"], start=sim_row + 2):
        ws.cell(row=j, column=1, value=s["threshold_kg"])
        ws.cell(row=j, column=2, value=s["below_count"])
        ws.cell(row=j, column=3, value=round(s["below_pct"], 1))
        ws.cell(row=j, column=4, value=round(s["below_total_kg"], 1))
        ws.cell(row=j, column=5, value=round(s["below_pct_of_total_kg"], 1))
        ws.cell(row=j, column=6, value=round(s["foregone_rev_rs"], 0))
        ws.cell(row=j, column=7, value=round(s["foregone_pct_of_total_rev"], 1))

    _autofit(ws)


def _write_cust_totals(ws, data):
    ws.cell(row=1, column=1, value="Auction-wide per-customer totals (sum of standalone per-lot allocations). May exceed monthly demand if standalone over-attributes.").font = Font(italic=True, color="888888")
    _write_header(ws, 3, ["customer", "total_kg", "monthly_demand_kg",
                          "fulfilment_%", "transport_bracket",
                          "revenue_at_rate_₹"])
    for i, r in enumerate(data["cust_totals"], start=4):
        ws.cell(row=i, column=1, value=r["customer"])
        ws.cell(row=i, column=2, value=round(r["total_kg"], 1))
        ws.cell(row=i, column=3, value=round(r["monthly_demand_kg"], 1))
        pct = (100 * r["total_kg"] / r["monthly_demand_kg"]
               if r["monthly_demand_kg"] else None)
        ws.cell(row=i, column=4, value=round(pct, 1) if pct is not None else "—")
        cell = ws.cell(row=i, column=5, value=r["bracket"])
        if r["bracket"] in ("<=5T",):
            cell.fill = BAD_FILL
        elif r["bracket"] == "5-10T":
            cell.fill = WARN_FILL
        ws.cell(row=i, column=6, value=round(r["revenue_at_rate"], 0))

    # Bracket summary
    summary_row = len(data["cust_totals"]) + 6
    _write_subheader(ws, summary_row, "Transport bracket summary", span=3)
    _write_header(ws, summary_row + 1, ["bracket", "customer_count", "total_kg"])
    bracket_data = defaultdict(lambda: {"count": 0, "kg": 0.0})
    for r in data["cust_totals"]:
        if r["total_kg"] > 0:
            b = r["bracket"]
            bracket_data[b]["count"] += 1
            bracket_data[b]["kg"] += r["total_kg"]
    for j, label in enumerate(TRANSPORT_LABELS, start=summary_row + 2):
        d = bracket_data.get(label, {"count": 0, "kg": 0})
        ws.cell(row=j, column=1, value=label)
        ws.cell(row=j, column=2, value=d["count"])
        ws.cell(row=j, column=3, value=round(d["kg"], 1))

    # Per-customer-per-auction MIN_QTY simulation
    sim_row = summary_row + 2 + len(TRANSPORT_LABELS) + 2
    _write_subheader(ws, sim_row,
                     "MIN_QTY simulation — customers BELOW threshold "
                     "(auction-wide totals)", span=7)
    _write_header(ws, sim_row + 1, ["MIN_QTY_kg", "below_count", "below_%",
                                    "below_total_kg", "below_%_of_total_kg",
                                    "foregone_rev_₹", "foregone_%_of_rev"])
    for j, s in enumerate(data["min_qty_sim_cust"], start=sim_row + 2):
        ws.cell(row=j, column=1, value=s["threshold_kg"])
        ws.cell(row=j, column=2, value=s["below_count"])
        ws.cell(row=j, column=3, value=round(s["below_pct"], 1))
        ws.cell(row=j, column=4, value=round(s["below_total_kg"], 1))
        ws.cell(row=j, column=5, value=round(s["below_pct_of_total_kg"], 1))
        ws.cell(row=j, column=6, value=round(s["foregone_rev_rs"], 0))
        ws.cell(row=j, column=7, value=round(s["foregone_pct_of_total_rev"], 1))

    _autofit(ws)


def _write_width_bands(ws, data):
    # Aggregate by band across lots
    agg: dict[str, dict] = {lab: {"count": 0, "kg": 0.0,
                                   "slit_count": 0, "slit_kg": 0.0,
                                   "asis_count": 0, "asis_kg": 0.0,
                                   "salv_count": 0, "salv_kg": 0.0}
                             for lab in WIDTH_LABELS}
    for r in data["lot_records"]:
        if not r["feasible"]:
            continue
        for band, v in r["by_band"].items():
            for k in agg[band]:
                agg[band][k] += v[k]

    ws.cell(row=1, column=1, value="Coil width-band breakdown + chosen disposition by band").font = Font(italic=True, color="888888")
    _write_header(ws, 3, ["band", "coil_count", "kg_total",
                          "slit_count", "slit_kg",
                          "asis_count", "asis_kg",
                          "salv_count", "salv_kg"])
    for i, lab in enumerate(WIDTH_LABELS, start=4):
        d = agg[lab]
        ws.cell(row=i, column=1, value=lab)
        ws.cell(row=i, column=2, value=d["count"])
        ws.cell(row=i, column=3, value=round(d["kg"], 1))
        ws.cell(row=i, column=4, value=d["slit_count"])
        ws.cell(row=i, column=5, value=round(d["slit_kg"], 1))
        ws.cell(row=i, column=6, value=d["asis_count"])
        ws.cell(row=i, column=7, value=round(d["asis_kg"], 1))
        ws.cell(row=i, column=8, value=d["salv_count"])
        ws.cell(row=i, column=9, value=round(d["salv_kg"], 1))

    # Headline note
    note_row = 4 + len(WIDTH_LABELS) + 2
    wide_slit_count = agg[">1300mm"]["slit_count"] + agg["1001-1300mm"]["slit_count"]
    wide_slit_kg = agg[">1300mm"]["slit_kg"] + agg["1001-1300mm"]["slit_kg"]
    ws.cell(row=note_row, column=1,
            value=f"NOTE: {wide_slit_count} coil(s) >1000mm chosen for slit, "
                  f"totalling {wide_slit_kg:,.0f} kg. "
                  f"Actual 'forbid slit on full coils' impact requires a re-solve "
                  f"(Tier 2).").font = Font(italic=True)
    _autofit(ws)


def _write_salvage(ws, data):
    rows = [(r["lot"], s) for r in data["lot_records"] if r["feasible"]
            for s in r["salvaged"]]
    ws.cell(row=1, column=1, value="Coils chosen for salvage (whole-coil at lot_start − coating delta)").font = Font(italic=True, color="888888")
    _write_header(ws, 3, ["lot", "batch", "coating", "width_mm",
                          "weight_kg", "lot_start_₹/kg", "delta_₹",
                          "salv_price_₹/kg", "salv_rev_₹",
                          "scrap_counterfact_₹", "savings_₹"])
    for i, (lot, s) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=lot)
        ws.cell(row=i, column=2, value=s["batch"])
        ws.cell(row=i, column=3, value=s["coating"])
        ws.cell(row=i, column=4, value=round(s["width_mm"], 1))
        ws.cell(row=i, column=5, value=round(s["weight_kg"], 1))
        ws.cell(row=i, column=6, value=round(s["lot_start_per_kg"], 2))
        ws.cell(row=i, column=7, value=round(s["delta"], 2))
        ws.cell(row=i, column=8, value=round(s["salv_price_per_kg"], 2))
        ws.cell(row=i, column=9, value=round(s["salv_rev_rs"], 0))
        ws.cell(row=i, column=10, value=round(s["scrap_counterfactual_rs"], 0))
        cell = ws.cell(row=i, column=11, value=round(s["savings_rs"], 0))
        if s["savings_rs"] > 0:
            cell.fill = GOOD_FILL
    # Per-coating frequency
    if rows:
        coats = defaultdict(lambda: {"count": 0, "kg": 0.0, "savings": 0.0})
        for _, s in rows:
            coats[s["coating"]]["count"] += 1
            coats[s["coating"]]["kg"] += s["weight_kg"]
            coats[s["coating"]]["savings"] += s["savings_rs"]
        summary_row = 4 + len(rows) + 2
        _write_subheader(ws, summary_row, "Per-coating salvage summary", span=4)
        _write_header(ws, summary_row + 1, ["coating", "count", "total_kg",
                                            "total_savings_₹"])
        for j, (c, d) in enumerate(sorted(coats.items()),
                                    start=summary_row + 2):
            ws.cell(row=j, column=1, value=c)
            ws.cell(row=j, column=2, value=d["count"])
            ws.cell(row=j, column=3, value=round(d["kg"], 1))
            ws.cell(row=j, column=4, value=round(d["savings"], 0))
    _autofit(ws)


def _write_max_bid_profit(ws, data):
    """Per-lot net profitability AT MAX-BID, after transport. Net margin must
    clear 2% to be worth pursuing.

    At max-bid, gross profit = 8% × revenue (engine uses 8% target margin).
    Subtracting transport → net margin = 8% − (transport % of revenue).
    Pass-2%-check ⇒ transport ≤ 6% of revenue.

    Transport is attributed per lot: customer's per-lot allocation × rate
    chosen by their AUCTION-WIDE total (truck size set by total flow).
    """
    cust_totals_by_name = {r["customer"]: r["total_kg"]
                           for r in data["cust_totals"]}
    orders = data["orders"]

    ws.cell(row=1, column=1, value="Per-lot profitability at MAX-BID after transport — net margin floor 2% (i.e. transport ≤ 6% of revenue).").font = Font(italic=True, color="888888")
    ws.cell(row=2, column=1, value="Transport bracket selected by each customer's AUCTION-WIDE total kg; rate × per-lot kg attributes the cost to lots.").font = Font(italic=True, color="888888")

    _write_header(ws, 4, ["lot", "max_bid_₹/kg", "weight_kg",
                          "bid_cost_₹", "revenue_₹", "slit_cost_₹",
                          "gross_profit@maxbid_₹", "gross_margin_%",
                          "transport_₹", "transport_%_of_rev",
                          "net_profit@maxbid_₹", "net_margin_%",
                          "passes_2%"])

    sum_bid = sum_rev = sum_slit = sum_transport = sum_net = 0.0
    feas_rows = 0
    for i, r in enumerate(data["lot_records"], start=5):
        ws.cell(row=i, column=1, value=r["lot"])
        if not r["feasible"]:
            ws.cell(row=i, column=12, value="ERROR")
            continue
        m = r["metrics"]
        wt_kg = m["total_wt"]
        max_bid = m["max_bid"]
        revenue = m["total_rev"]
        slit_cost = r.get("slit_cost_rs", 0.0)
        bid_cost = max_bid * wt_kg
        gross_profit = revenue - bid_cost - slit_cost
        gross_margin = (100 * gross_profit / revenue) if revenue else 0
        transport = _lot_transport_cost(r, orders, cust_totals_by_name)
        transport_pct = (100 * transport / revenue) if revenue else 0
        net_profit = gross_profit - transport
        net_margin = (100 * net_profit / revenue) if revenue else 0
        passes = net_margin >= 2.0

        ws.cell(row=i, column=2, value=round(max_bid, 2))
        ws.cell(row=i, column=3, value=round(wt_kg, 1))
        ws.cell(row=i, column=4, value=round(bid_cost, 0))
        ws.cell(row=i, column=5, value=round(revenue, 0))
        ws.cell(row=i, column=6, value=round(slit_cost, 0))
        ws.cell(row=i, column=7, value=round(gross_profit, 0))
        ws.cell(row=i, column=8, value=round(gross_margin, 2))
        ws.cell(row=i, column=9, value=round(transport, 0))
        ws.cell(row=i, column=10, value=round(transport_pct, 2))
        ws.cell(row=i, column=11, value=round(net_profit, 0))
        nm_cell = ws.cell(row=i, column=12, value=round(net_margin, 2))
        flag_cell = ws.cell(row=i, column=13, value="YES" if passes else "NO")
        if passes:
            nm_cell.fill = GOOD_FILL
            flag_cell.fill = GOOD_FILL
        else:
            nm_cell.fill = BAD_FILL
            flag_cell.fill = BAD_FILL

        sum_bid += bid_cost
        sum_rev += revenue
        sum_slit += slit_cost
        sum_transport += transport
        sum_net += net_profit
        feas_rows += 1

    # Aggregate row
    if feas_rows:
        agg_row = 5 + len(data["lot_records"]) + 1
        agg_net_margin = (100 * sum_net / sum_rev) if sum_rev else 0
        ws.cell(row=agg_row, column=1, value="TOTAL (all lots)").font = SUB_FONT
        ws.cell(row=agg_row, column=4, value=round(sum_bid, 0)).font = SUB_FONT
        ws.cell(row=agg_row, column=5, value=round(sum_rev, 0)).font = SUB_FONT
        ws.cell(row=agg_row, column=6, value=round(sum_slit, 0)).font = SUB_FONT
        ws.cell(row=agg_row, column=9, value=round(sum_transport, 0)).font = SUB_FONT
        ws.cell(row=agg_row, column=11, value=round(sum_net, 0)).font = SUB_FONT
        agg_cell = ws.cell(row=agg_row, column=12, value=round(agg_net_margin, 2))
        agg_cell.font = SUB_FONT
        agg_cell.fill = GOOD_FILL if agg_net_margin >= 2.0 else BAD_FILL

        # BIDABLE-only total — exclude lots where max-bid < lot_start (the model says don't bid)
        bid_bid = bid_rev = bid_slit = bid_transport = bid_net = 0.0
        bid_count = 0
        for r in data["lot_records"]:
            if not r["feasible"]:
                continue
            m = r["metrics"]
            start = r["coils"][0]["price_per_kg"] if r["coils"] else 0
            if m["max_bid"] < start:
                continue
            wt_kg = m["total_wt"]
            revenue = m["total_rev"]
            slit_cost = r.get("slit_cost_rs", 0.0)
            bid_cost = m["max_bid"] * wt_kg
            transport = _lot_transport_cost(r, orders, cust_totals_by_name)
            net_profit = revenue - bid_cost - slit_cost - transport
            bid_bid += bid_cost
            bid_rev += revenue
            bid_slit += slit_cost
            bid_transport += transport
            bid_net += net_profit
            bid_count += 1

        if bid_count:
            bid_row = agg_row + 1
            bid_net_margin = (100 * bid_net / bid_rev) if bid_rev else 0
            ws.cell(row=bid_row, column=1,
                    value=f"TOTAL bidable only ({bid_count} of {feas_rows} lots)").font = SUB_FONT
            ws.cell(row=bid_row, column=4, value=round(bid_bid, 0)).font = SUB_FONT
            ws.cell(row=bid_row, column=5, value=round(bid_rev, 0)).font = SUB_FONT
            ws.cell(row=bid_row, column=6, value=round(bid_slit, 0)).font = SUB_FONT
            ws.cell(row=bid_row, column=9, value=round(bid_transport, 0)).font = SUB_FONT
            ws.cell(row=bid_row, column=11, value=round(bid_net, 0)).font = SUB_FONT
            bm_cell = ws.cell(row=bid_row, column=12, value=round(bid_net_margin, 2))
            bm_cell.font = SUB_FONT
            bm_cell.fill = GOOD_FILL if bid_net_margin >= 2.0 else BAD_FILL

    _autofit(ws)


def _write_net_profit(ws, data):
    """Per customer: gross customer-rev, transport cost (per bracket), net."""
    rows = []
    for r in data["cust_totals"]:
        if r["total_kg"] <= 0:
            continue
        cost, label = _transport_cost(r["customer"], r["total_kg"])
        rows.append({
            "customer": r["customer"],
            "total_kg": r["total_kg"],
            "gross_rev_rs": r["revenue_at_rate"],
            "transport_cost_rs": cost,
            "net_rev_rs": r["revenue_at_rate"] - cost,
            "transport_pct_of_gross": (100 * cost / r["revenue_at_rate"]
                                        if r["revenue_at_rate"] else 0),
            "bracket": label,
        })

    ws.cell(row=1, column=1, value="Per-customer net profitability — gross customer revenue minus transport cost (per-tonne by bracket).").font = Font(italic=True, color="888888")
    ws.cell(row=2, column=1, value="Assumes rate × actual_kg (literal 'per tonne'). If brackets are full-truck flat rates, scale accordingly.").font = Font(italic=True, color="888888")

    _write_header(ws, 4, ["customer", "total_kg", "gross_rev_₹",
                          "transport_cost_₹", "transport_%_of_gross",
                          "net_rev_₹", "bracket_used"])
    for i, r in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=r["customer"])
        ws.cell(row=i, column=2, value=round(r["total_kg"], 1))
        ws.cell(row=i, column=3, value=round(r["gross_rev_rs"], 0))
        ws.cell(row=i, column=4, value=round(r["transport_cost_rs"], 0))
        cell = ws.cell(row=i, column=5,
                       value=round(r["transport_pct_of_gross"], 2))
        if r["transport_pct_of_gross"] >= 5:
            cell.fill = BAD_FILL
        elif r["transport_pct_of_gross"] >= 2:
            cell.fill = WARN_FILL
        ws.cell(row=i, column=6, value=round(r["net_rev_rs"], 0))
        ws.cell(row=i, column=7, value=r["bracket"])

    # Totals row
    tot_kg = sum(r["total_kg"] for r in rows)
    tot_gross = sum(r["gross_rev_rs"] for r in rows)
    tot_cost = sum(r["transport_cost_rs"] for r in rows)
    tot_net = tot_gross - tot_cost
    tot_pct = (100 * tot_cost / tot_gross) if tot_gross else 0
    tot_row = 5 + len(rows) + 1
    ws.cell(row=tot_row, column=1, value="TOTAL").font = SUB_FONT
    ws.cell(row=tot_row, column=2, value=round(tot_kg, 1)).font = SUB_FONT
    ws.cell(row=tot_row, column=3, value=round(tot_gross, 0)).font = SUB_FONT
    ws.cell(row=tot_row, column=4, value=round(tot_cost, 0)).font = SUB_FONT
    ws.cell(row=tot_row, column=5, value=round(tot_pct, 2)).font = SUB_FONT
    ws.cell(row=tot_row, column=6, value=round(tot_net, 0)).font = SUB_FONT
    _autofit(ws)


def _write_flags(ws, data):
    ws.cell(row=1, column=1, value="Data-hygiene flags (parse anomalies + safeguards)").font = Font(italic=True, color="888888")
    _write_header(ws, 3, ["#", "flag"])
    for i, f in enumerate(data["flags"], start=4):
        ws.cell(row=i, column=1, value=i - 3)
        cell = ws.cell(row=i, column=2, value=f)
        if "salvage-hygiene" in f or "demand-mismatch" in f:
            cell.fill = WARN_FILL
    _autofit(ws)


def write_auction_workbook(auction_name, data, output_path):
    wb = Workbook()
    ws = wb.active
    _write_summary(ws, auction_name, data)
    _write_per_lot_summary(wb.create_sheet("Per-Lot Summary"), data)
    _write_lot_pnl_detail(wb.create_sheet("Lot P&L by Tier"), data)
    _write_lot_detail(wb.create_sheet("Lot Detail (Cut Sheets)"), data)
    _write_customer_fulfillment(
        wb.create_sheet("Customer Fulfillment"), data)
    _write_cust_lot_alloc(wb.create_sheet("Customer × Lot"), data)
    _write_bidable_summary(wb.create_sheet("Bidable Auction Summary"), data)
    _write_salvage(wb.create_sheet("Salvage"), data)
    _write_width_bands(wb.create_sheet("Width Bands"), data)
    _write_flags(wb.create_sheet("Flags"), data)
    wb.save(output_path)


# ── cross-auction summary ────────────────────────────────────────────────────

def write_cross_auction_summary(per_auction, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    tiers = _get_bid_tiers()
    primary_name, primary_margin = tiers[0]
    headers = ["auction", "coils", "lots", "total_kg",
               "profit_₹", f"avg_{primary_name}_bid_₹/kg", "avg_margin_%",
               "scrap_kg", "scrap_%",
               "salvage_count", "salvage_savings_₹",
               "flags", "total_solve_s"]
    _write_header(ws, 1, headers)
    for i, (name, data) in enumerate(per_auction, start=2):
        feas = [r for r in data["lot_records"] if r["feasible"]]
        total_kg = sum(c["weight_g"] for c in data["coils"]) / 1000
        profit = sum(r["metrics"]["profit"] for r in feas)
        cust_by_name = {r["customer"]: r["total_kg"]
                        for r in data["cust_totals"]}
        bid_w = sum(
            _bid_for_net_margin(r["metrics"]["total_rev"],
                                r.get("slit_cost_rs", 0),
                                _lot_transport_cost(r, data["orders"],
                                                    cust_by_name),
                                r["metrics"]["total_wt"], primary_margin)
            * r["metrics"]["total_wt"]
            for r in feas
        )
        bid_kg = sum(r["metrics"]["total_wt"] for r in feas)
        avg_bid = bid_w / bid_kg if bid_kg else 0
        rev = sum(r["metrics"]["total_rev"] for r in feas)
        margin = (100 * profit / rev) if rev else 0
        scrap = sum(r["kg_scrap"] for r in feas)
        salv_n = sum(len(r["salvaged"]) for r in feas)
        salv_s = sum(s["savings_rs"]
                     for r in feas for s in r["salvaged"])
        solve_s = sum(r["solve_time_s"] for r in data["lot_records"])
        row = [name, len(data["coils"]), len(data["lot_records"]),
               round(total_kg, 0), round(profit, 0),
               round(avg_bid, 2), round(margin, 2),
               round(scrap, 0),
               round(100 * scrap / total_kg, 2) if total_kg else 0,
               salv_n, round(salv_s, 0),
               len(data["flags"]), round(solve_s, 1)]
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)
    _autofit(ws)

    # Aggregated MIN_QTY simulation across all auctions
    ws2 = wb.create_sheet("Aggregated MIN_QTY")
    all_alloc = [r for _, d in per_auction for r in d["cust_lot_alloc"]]
    total_count = len(all_alloc)
    total_kg = sum(r["kg"] for r in all_alloc)
    total_rev = sum(r["revenue_at_rate"] for r in all_alloc)
    ws2.cell(row=1, column=1, value=f"Pooled across {len(per_auction)} auctions — "
                                    f"{total_count} (customer, lot) allocations, "
                                    f"{total_kg:,.0f} kg, ₹{total_rev:,.0f} revenue").font = Font(italic=True, color="888888")
    _write_header(ws2, 3, ["MIN_QTY_kg", "below_count", "below_%",
                           "below_total_kg", "below_%_of_total_kg",
                           "foregone_rev_₹", "foregone_%_of_rev"])
    for j, thr in enumerate(MIN_QTY_THRESHOLDS_KG, start=4):
        below = [r for r in all_alloc if r["kg"] < thr]
        ws2.cell(row=j, column=1, value=thr)
        ws2.cell(row=j, column=2, value=len(below))
        ws2.cell(row=j, column=3, value=round(100 * len(below) / total_count, 1)
                 if total_count else 0)
        ws2.cell(row=j, column=4, value=round(sum(r["kg"] for r in below), 1))
        ws2.cell(row=j, column=5,
                 value=round(100 * sum(r["kg"] for r in below) / total_kg, 1)
                 if total_kg else 0)
        ws2.cell(row=j, column=6,
                 value=round(sum(r["revenue_at_rate"] for r in below), 0))
        ws2.cell(row=j, column=7,
                 value=round(100 * sum(r["revenue_at_rate"] for r in below)
                             / total_rev, 1)
                 if total_rev else 0)
    _autofit(ws2)

    # Max-bid profitability across auctions (after transport, 2% floor check)
    ws3 = wb.create_sheet("Max-Bid Profitability")
    ws3.cell(row=1, column=1, value=f"Per-auction net profitability at PRIMARY bid ({primary_name}, {int(primary_margin*100)}% net) after transport. Net margin ≥ 2% = pursue.").font = Font(italic=True, color="888888")
    _write_header(ws3, 3, ["auction", "lots", "bidable_lots",
                           "bid_cost_₹", "revenue_₹", "slit_cost_₹",
                           "transport_₹", "net_profit_₹", "net_margin_%",
                           "bidable_net_profit_₹", "bidable_net_margin_%",
                           "passes_2%"])
    grand = {"bid": 0.0, "rev": 0.0, "slit": 0.0, "tr": 0.0, "net": 0.0,
             "b_bid": 0.0, "b_rev": 0.0, "b_slit": 0.0, "b_tr": 0.0,
             "b_net": 0.0, "lots": 0, "b_lots": 0}
    for i, (name, data) in enumerate(per_auction, start=4):
        cust_by_name = {r["customer"]: r["total_kg"]
                         for r in data["cust_totals"]}
        sum_b = sum_r = sum_sl = sum_t = sum_n = 0.0
        b_b = b_r = b_sl = b_t = b_n = 0.0
        b_lots = 0
        all_lots = 0
        for r in data["lot_records"]:
            if not r["feasible"]:
                continue
            all_lots += 1
            m = r["metrics"]
            wt_kg = m["total_wt"]
            revenue = m["total_rev"]
            slit_cost = r.get("slit_cost_rs", 0.0)
            transport = _lot_transport_cost(r, data["orders"], cust_by_name)
            primary_bid = _bid_for_net_margin(revenue, slit_cost, transport,
                                              wt_kg, primary_margin)
            bid_cost = primary_bid * wt_kg
            net = revenue - bid_cost - slit_cost - transport
            sum_b += bid_cost
            sum_r += revenue
            sum_sl += slit_cost
            sum_t += transport
            sum_n += net
            start = r["coils"][0]["price_per_kg"] if r["coils"] else 0
            if primary_bid >= start:
                b_b += bid_cost
                b_r += revenue
                b_sl += slit_cost
                b_t += transport
                b_n += net
                b_lots += 1
        net_margin = (100 * sum_n / sum_r) if sum_r else 0
        b_margin = (100 * b_n / b_r) if b_r else 0
        ws3.cell(row=i, column=1, value=name)
        ws3.cell(row=i, column=2, value=all_lots)
        ws3.cell(row=i, column=3, value=b_lots)
        ws3.cell(row=i, column=4, value=round(sum_b, 0))
        ws3.cell(row=i, column=5, value=round(sum_r, 0))
        ws3.cell(row=i, column=6, value=round(sum_sl, 0))
        ws3.cell(row=i, column=7, value=round(sum_t, 0))
        ws3.cell(row=i, column=8, value=round(sum_n, 0))
        nm = ws3.cell(row=i, column=9, value=round(net_margin, 2))
        nm.fill = GOOD_FILL if net_margin >= 2 else BAD_FILL
        ws3.cell(row=i, column=10, value=round(b_n, 0))
        bm = ws3.cell(row=i, column=11, value=round(b_margin, 2))
        bm.fill = GOOD_FILL if b_margin >= 2 else BAD_FILL
        ws3.cell(row=i, column=12, value="YES" if b_margin >= 2 else "NO")
        grand["bid"] += sum_b; grand["rev"] += sum_r; grand["slit"] += sum_sl
        grand["tr"] += sum_t; grand["net"] += sum_n
        grand["b_bid"] += b_b; grand["b_rev"] += b_r; grand["b_slit"] += b_sl
        grand["b_tr"] += b_t; grand["b_net"] += b_n
        grand["lots"] += all_lots; grand["b_lots"] += b_lots
    # Grand totals
    grand_row = 4 + len(per_auction) + 1
    g_margin = (100 * grand["net"] / grand["rev"]) if grand["rev"] else 0
    bg_margin = (100 * grand["b_net"] / grand["b_rev"]) if grand["b_rev"] else 0
    ws3.cell(row=grand_row, column=1, value="GRAND TOTAL").font = SUB_FONT
    ws3.cell(row=grand_row, column=2, value=grand["lots"]).font = SUB_FONT
    ws3.cell(row=grand_row, column=3, value=grand["b_lots"]).font = SUB_FONT
    ws3.cell(row=grand_row, column=4, value=round(grand["bid"], 0)).font = SUB_FONT
    ws3.cell(row=grand_row, column=5, value=round(grand["rev"], 0)).font = SUB_FONT
    ws3.cell(row=grand_row, column=6, value=round(grand["slit"], 0)).font = SUB_FONT
    ws3.cell(row=grand_row, column=7, value=round(grand["tr"], 0)).font = SUB_FONT
    ws3.cell(row=grand_row, column=8, value=round(grand["net"], 0)).font = SUB_FONT
    gm = ws3.cell(row=grand_row, column=9, value=round(g_margin, 2))
    gm.font = SUB_FONT
    gm.fill = GOOD_FILL if g_margin >= 2 else BAD_FILL
    ws3.cell(row=grand_row, column=10, value=round(grand["b_net"], 0)).font = SUB_FONT
    bgm = ws3.cell(row=grand_row, column=11, value=round(bg_margin, 2))
    bgm.font = SUB_FONT
    bgm.fill = GOOD_FILL if bg_margin >= 2 else BAD_FILL
    ws3.cell(row=grand_row, column=12,
             value="YES" if bg_margin >= 2 else "NO").font = SUB_FONT
    _autofit(ws3)

    wb.save(output_path)


# ── driver ───────────────────────────────────────────────────────────────────

def process_auction(auction_path, orders, time_limit_s):
    print(f"\n── {auction_path.name} ──")
    coils, auction_flags = ar.parse_auction(auction_path)
    print(f"  parsed: {len(coils)} coils, {len({c['lot'] for c in coils})} lots")
    by_lot: dict[str, list[dict]] = defaultdict(list)
    for c in coils:
        by_lot[c["lot"]].append(c)
    lot_records = []
    for lot in sorted(by_lot):
        print(f"  solving lot {lot} …", end=" ", flush=True)
        r = _build_lot_record(lot, by_lot[lot], orders, time_limit_s)
        if r["feasible"]:
            print(f"max-bid ₹{r['metrics']['max_bid']:,.2f}/kg, "
                  f"profit ₹{r['metrics']['profit']:,.0f}, "
                  f"{r['solve_time_s']:.1f}s")
        else:
            print(f"INFEASIBLE ({r['status']}) in {r['solve_time_s']:.1f}s")
        lot_records.append(r)
    return {
        "coils": coils,
        "orders": orders,
        "n_customers": len({o["customer"] for o in orders}),
        "lot_records": lot_records,
        "auction_flags": auction_flags,
    }


def parse_args():
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0] if __doc__ else "")
    p.add_argument("--customers", required=True,
                   help="Path to the customer workbook (.xlsx)")
    p.add_argument("--auctions-dir", required=True,
                   help="Folder containing auction .xlsx files")
    p.add_argument("--only", default=None,
                   help="Optional substring; run only auction files whose "
                        "name contains it (case-insensitive). e.g. '4.4'")
    p.add_argument("--output-dir", required=True,
                   help="Where to write the reports")
    p.add_argument("--time-limit-s", type=int, default=60,
                   help="Per-lot CP-SAT time budget, seconds (default 60)")
    return p.parse_args()


def main():
    args = parse_args()
    # Monkey-patch the engine's import-time time-limit (it's read at solve time
    # via the module-level binding).
    eng.SOLVE_TIME_LIMIT_S = args.time_limit_s
    print(f"engine time limit: {eng.SOLVE_TIME_LIMIT_S}s/lot")
    print(f"engine band       : {os.environ.get('SLIT_BAND_MM', '650 (default)')}")
    print(f"engine salvage    : {os.environ.get('SLIT_SALVAGE', '(empty)')}")

    customers_path = Path(args.customers)
    auctions_dir = Path(args.auctions_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\nparsing customers: {customers_path.name}")
    orders, customer_flags = cr.parse_customers(customers_path)
    print(f"  {len(orders)} orders across "
          f"{len({o['customer'] for o in orders})} customers")
    if customer_flags:
        print(f"  {len(customer_flags)} parse flags")

    auction_files = sorted(p for p in auctions_dir.glob("*.xlsx")
                           if not p.name.startswith("~$"))
    if args.only:
        needle = args.only.lower()
        auction_files = [p for p in auction_files if needle in p.name.lower()]
        print(f"filter '--only {args.only}' matched {len(auction_files)} file(s)")
    if not auction_files:
        print(f"no auctions found in {auctions_dir}"
              + (f" matching '--only {args.only}'" if args.only else ""))
        return 1
    print(f"\nauctions found: {len(auction_files)}")

    per_auction = []
    for ap in auction_files:
        data = process_auction(ap, orders, args.time_limit_s)
        # Aggregations
        data["cust_lot_alloc"] = _customer_lot_alloc(data["lot_records"],
                                                     orders)
        data["cust_totals"] = _customer_totals(data["cust_lot_alloc"], orders)
        data["min_qty_sim"] = _min_qty_sim(data["cust_lot_alloc"], "kg")
        data["min_qty_sim_cust"] = _min_qty_sim(data["cust_totals"],
                                                "total_kg")
        data["flags"] = _flags(data["lot_records"], data["coils"], orders,
                                data["auction_flags"], customer_flags)
        # Write per-auction workbook
        safe_name = ap.stem.replace(" ", "_").replace("(", "").replace(")", "")
        out = output_dir / f"measurement_{safe_name}.xlsx"
        write_auction_workbook(ap.stem, data, out)
        print(f"  → wrote {out.relative_to(output_dir.parent)}")
        per_auction.append((ap.stem, data))

    summary_out = output_dir / "cross_auction_summary.xlsx"
    write_cross_auction_summary(per_auction, summary_out)
    print(f"\n→ wrote {summary_out.relative_to(output_dir.parent)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
