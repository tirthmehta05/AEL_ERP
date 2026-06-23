"""One-time helper: convert customer-grade cells from grade-only form
(e.g., "600|530") to the exact Thickness+C+Grade form used by auction files
(e.g., "50c600|50c530"). Writes a NEW workbook so the original is preserved.

This is the migration step that goes with the parser change to exact-string
matching (lower-case normalized). Future grades that need different chemistry
(e.g., CRGO "50CR600") will then be unambiguously distinct from CRNO "50C600".

Logic per (Grades, Thickness) cell pair:
  - Split grades on "|"; strip + lowercase each
  - If a grade already matches /^\\d+c\\d+$/  → keep as-is (already in target form)
  - If a grade matches /^\\d+$/ (bare number) → expand using the row's thickness
        column: for each thickness t in that row → "<int(t*100)>c<grade>".
        Row with thickness "0.5|0.6" + grade "600" becomes "50c600|60c600".
  - Anything else → keep normalized + emit warning.

Usage:
    uv run python -m tools.migrate_customer_grades \\
        --input data/sample_customers_v2.xlsx \\
        --output data/sample_customers_v3.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

GRADE_FULL_RE = re.compile(r"^\d+c\d+$")
GRADE_BARE_RE = re.compile(r"^\d+(?:\.0+)?$")  # accepts "800" or "800.0"


def _split_pipe(v) -> list[str]:
    if v is None or (isinstance(v, float) and v != v):  # NaN
        return []
    return [p.strip() for p in str(v).split("|") if p.strip()]


def _migrate_grade_cell(grades_val, thickness_val) -> tuple[str | None, list[str]]:
    """Return (new_grades_string_or_None, warnings).
    new_grades is None when the cell was empty (preserve)."""
    grades = _split_pipe(grades_val)
    if not grades:
        return None, []
    thks = _split_pipe(thickness_val)
    warnings: list[str] = []
    out: list[str] = []
    for g in grades:
        g_norm = g.lower()
        if GRADE_FULL_RE.match(g_norm):
            out.append(g_norm)
        elif GRADE_BARE_RE.match(g_norm):
            # Strip a trailing ".0" (pandas-read-as-float artifact); keep
            # actual decimal grades flagged below.
            g_int = g_norm.split(".")[0]
            if not thks:
                warnings.append(
                    f"bare grade '{g}' but no thickness in row — kept as-is")
                out.append(g_int)
                continue
            for t in thks:
                try:
                    tf = float(t)
                    prefix = f"{int(round(tf * 100))}c"
                    out.append(f"{prefix}{g_int}")
                except (ValueError, TypeError):
                    warnings.append(
                        f"thickness '{t}' not numeric — grade '{g}' kept bare")
                    out.append(g_int)
        else:
            warnings.append(
                f"grade '{g}' unrecognized form — kept normalized")
            out.append(g_norm)
    return "|".join(out), warnings


def _find_col(ws, *keywords) -> int | None:
    for cell in ws[1]:
        if cell.value is None:
            continue
        h = str(cell.value).strip().lower()
        if any(kw in h for kw in keywords):
            return cell.column
    return None


def migrate(input_path: Path, output_path: Path) -> tuple[int, list[str]]:
    wb = load_workbook(input_path)
    total_changes = 0
    warnings: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        grade_col = _find_col(ws, "grade")
        thk_col = _find_col(ws, "thickness", "thk")
        if grade_col is None or thk_col is None:
            warnings.append(
                f"[{sheet_name}] missing grade or thickness column — skipped")
            continue

        for row_idx in range(2, ws.max_row + 1):
            grades_cell = ws.cell(row=row_idx, column=grade_col)
            thk_cell = ws.cell(row=row_idx, column=thk_col)
            old = grades_cell.value
            new, warns = _migrate_grade_cell(old, thk_cell.value)
            if new is not None and new != old:
                grades_cell.value = new
                total_changes += 1
            for w in warns:
                warnings.append(f"[{sheet_name} row {row_idx}] {w}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return total_changes, warnings


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument("--input", required=True, help="source .xlsx")
    p.add_argument("--output", required=True, help="destination .xlsx")
    args = p.parse_args()

    src = Path(args.input)
    if not src.exists():
        print(f"ERROR: {src} not found")
        return 1

    changes, warnings = migrate(src, Path(args.output))
    print(f"\nMigration complete.")
    print(f"  input : {src}")
    print(f"  output: {args.output}")
    print(f"  cells changed: {changes}")
    if warnings:
        print(f"\nWarnings ({len(warnings)}):")
        for w in warnings:
            print(f"  {w}")
    else:
        print(f"  no warnings")
    return 0


if __name__ == "__main__":
    sys.exit(main())
