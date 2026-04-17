from typing import List, Dict, Optional, Tuple
import json
import io
import pandas as pd
from fpdf import FPDF
from num2words import num2words
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.shared.utils.logger_config import setup_logger
from src.shared.integrations.google_drive_service import GoogleDriveService
from src.data_entry.models.delivery_challan_models import (
    DeliveryChallanRequest,
    DeliveryChallanRecord,
    DeliveryChallanItem,
    MissingFieldResult,
)

import config


class DeliveryChallanPDF(FPDF):
    def header(self):
        pass

    def footer(self):
        pass


class DeliveryChallanService:
    def __init__(self):
        self.logger = setup_logger(__name__)
        self.google_service = GoogleDriveService()
        self._consignees_cache = None
        self._rm_inward_cache = None

    # ──────────────────────────────────────────
    # Consignee Master
    # ──────────────────────────────────────────

    def get_consignees(self) -> List[Dict]:
        if self._consignees_cache is not None:
            return self._consignees_cache

        df = self.google_service.get_worksheet_data(
            config.settings.api.google_sheets_id,
            config.CONSIGNEE_MASTER_SHEET,
        )

        consignees = df.to_dict(orient="records")

        self._consignees_cache = consignees
        return consignees

    # ──────────────────────────────────────────
    # RM Inward Data Access
    # ──────────────────────────────────────────

    def get_rm_inward_data(self) -> pd.DataFrame:
        """Fetch RM Inward sheet data (header_row=3 to match existing convention)."""
        if self._rm_inward_cache is not None:
            return self._rm_inward_cache

        df = self.google_service.get_worksheet_data(
            config.settings.api.google_sheets_id,
            config.RM_INWARD_ISSUE_SHEET,
            header_row=3,
        )

        df.columns = df.columns.str.strip()
        self._rm_inward_cache = df
        return df

    def get_coil_numbers(self) -> List[str]:
        """Get all unique coil numbers from RM Inward sheet."""
        df = self.get_rm_inward_data()

        if "Coil No" not in df.columns:
            return []

        return (
            df["Coil No"]
            .dropna()
            .astype(str)
            .str.strip()
            .loc[lambda s: s != ""]
            .unique()
            .tolist()
        )

    def get_rm_parties(self) -> List[str]:
        """Get all unique supplier/party names from RM Inward sheet."""
        df = self.get_rm_inward_data()

        col = "Coil Supplier" if "Coil Supplier" in df.columns else "Supplier"
        if col not in df.columns:
            return []

        return sorted(
            df[col]
            .dropna()
            .astype(str)
            .str.strip()
            .loc[lambda s: s != ""]
            .unique()
            .tolist()
        )

    def get_coils_by_party(self, party: str) -> List[str]:
        """Get coil numbers filtered by a specific party/supplier."""
        df = self.get_rm_inward_data()

        col = "Coil Supplier" if "Coil Supplier" in df.columns else "Supplier"
        if col not in df.columns or "Coil No" not in df.columns:
            return []

        filtered = df[
            df[col].astype(str).str.strip().str.lower()
            == str(party).strip().lower()
        ]
        return (
            filtered["Coil No"]
            .dropna()
            .astype(str)
            .str.strip()
            .loc[lambda s: s != ""]
            .tolist()
        )

    def _coil_row_to_item(self, row: pd.Series) -> DeliveryChallanItem:
        """Convert one RM Inward row to a DeliveryChallanItem."""
        weight_kg = float(row.get("Coil Weight", 0) or 0)
        weight_mt = weight_kg / 1000

        return DeliveryChallanItem(
            coil_number=str(row.get("Coil No", "")),
            description=str(row.get("Grade", "")),
            grade=str(row.get("Grade", "")),
            width=float(row.get("Width", 0) or 0),
            thk=float(row.get("Thk", 0) or 0),
            coating=str(row.get("Coating", "")),
            weight_kg=weight_kg,
            weight_mt=weight_mt,
            nature_of_processing="Slitting",
            hsn_code="7208",
            rate_per_mt=0,
            value=0,
        )

    def _get_party_from_row(self, row: pd.Series) -> str:
        """Extract the supplier/party name from an RM Inward row."""
        col = "Coil Supplier" if "Coil Supplier" in row.index else "Supplier"
        return str(row.get(col, ""))

    def build_request_from_coils(
    self,
    coil_numbers: List[str],
    challan_type: str = "job_work",
    party_name: str = "",
    vehicle_no: str = "",
    ewaybill_no: str = "",
    hsn_code: str = "7208",
    rounded_off: float = 0.0,   
    remarks: str = "",
    ) -> DeliveryChallanRequest:
        """Build a DeliveryChallanRequest from RM Inward using remaining weight logic."""

        df = self.get_rm_inward_data()

        # --- Load RM Used once ---
        used_df = self.google_service.get_worksheet_data(
            config.settings.api.google_sheets_id,
            config.RM_USED_SHEET,
        )

        used_df.columns = [str(col).strip() for col in used_df.columns]

        used_df["Coil No"] = used_df["Coil No"].astype(str).str.strip()

        # detect correct weight column

        if "Coil Weight" in used_df.columns:
            used_weight_col = "Coil Weight"
        elif "Weight" in used_df.columns:
            used_weight_col = "Weight"
        elif "weight_kg" in used_df.columns:
            used_weight_col = "weight_kg"
        else:
            raise ValueError(f"No valid weight column in RM Used: {used_df.columns.tolist()}")

        used_df[used_weight_col] = pd.to_numeric(
            used_df[used_weight_col], errors="coerce"
        ).fillna(0)

        # aggregate used weight
        used_agg = (
            used_df.groupby("Coil No")[used_weight_col]
            .sum()
            .to_dict()
        )

        items = []
        detected_party = party_name

        for coil in coil_numbers:
            match = df[
                df["Coil No"].astype(str).str.strip()
                == str(coil).strip()
            ]

            if match.empty:
                raise ValueError(f"Coil '{coil}' not found in RM Inward sheet")

            row = match.iloc[0]

            inward_weight = float(row.get("Coil Weight", 0))
            used_weight = float(used_agg.get(str(coil).strip(), 0))

            remaining_weight = inward_weight - used_weight

            # skip fully used
            if remaining_weight <= 0:
                continue

            row = row.copy()
            row["Coil Weight"] = remaining_weight

            items.append(self._coil_row_to_item(row))

            if not detected_party:
                detected_party = self._get_party_from_row(row)

        if not items:
            raise ValueError("All selected coils are fully used.")

        return DeliveryChallanRequest(
        challan_type=challan_type,
        challan_date=datetime.now().date(),
        party_name=detected_party or "TAIIN STEEL FAB & INFRA PVT LTD",
        vehicle_no=vehicle_no or None,
        ewaybill_no=ewaybill_no or None,
        hsn_code=hsn_code,
        rounded_off=rounded_off,
        remarks=remarks or "Generated from RM Inward",
        items=items,
        )

    # ──────────────────────────────────────────
    # Unused Coils
    # ──────────────────────────────────────────

    def get_unused_coils(self):

        # --- Load RM Inward ---
        inward_df = self.google_service.get_worksheet_data(
        config.settings.api.google_sheets_id,
        config.RM_INWARD_ISSUE_SHEET,
        )
        inward_df.columns = [str(col).strip() for col in inward_df.columns]
        # --- Load RM Used ---
        used_df = self.google_service.get_worksheet_data(
        config.settings.api.google_sheets_id,
        config.RM_USED_SHEET,
            )
        used_df.columns = [str(col).strip() for col in used_df.columns]
        # --- Normalize column names ---
        inward_df["Coil No"] = inward_df["Coil No"].astype(str).str.strip()
        used_df["Coil No"] = used_df["Coil No"].astype(str).str.strip()
        # --- Handle weight columns safely ---
        if "Coil Weight" in inward_df.columns:
            inward_weight_col = "Coil Weight"
        elif "Weight" in inward_df.columns:
            inward_weight_col = "Weight"
        else:
            raise ValueError(f"No valid weight column in RM Inward: {inward_df.columns.tolist()}")

        inward_df[inward_weight_col] = pd.to_numeric(
            inward_df[inward_weight_col], errors="coerce"
        ).fillna(0)
        # IMPORTANT: Change this if RM Used column name is different
        used_weight_col = "Weight"
        if "Weight" not in used_df.columns:
            used_weight_col = "weight_kg"
        used_df[used_weight_col] = pd.to_numeric(used_df[used_weight_col], errors="coerce").fillna(0)
        # --- Aggregate RM Used ---
        used_agg = (
            used_df.groupby("Coil No")[used_weight_col]
            .sum()
            .reset_index()
            .rename(columns={used_weight_col: "Used Weight"})
        )
        # --- Merge inward + used ---
        merged = inward_df.merge(
            used_agg,
            on="Coil No",
            how="left"
        )
        merged["Used Weight"] = merged["Used Weight"].fillna(0)
        # --- Calculate remaining ---
        merged["Remaining Weight"] = merged[inward_weight_col] - merged["Used Weight"]
        # --- Keep only usable coils ---
        unused_df = merged[merged["Remaining Weight"] > 0].copy()
        # --- Replace weight with remaining ---
        unused_df[inward_weight_col] = unused_df["Remaining Weight"]
        # Optional: rename for UI clarity
        unused_df.rename(columns={"Weight": "Remaining Weight"}, inplace=True)

        return unused_df

    # ──────────────────────────────────────────
    # Missing Data Handling
    # ──────────────────────────────────────────

    def check_missing_fields(
        self, request: DeliveryChallanRequest
    ) -> List[MissingFieldResult]:
        """Check a challan request for missing critical data needed for PDF generation."""
        missing = []

        # Check user-input fields
        if not request.vehicle_no or not request.vehicle_no.strip():
            missing.append(MissingFieldResult(
                field_name="vehicle_no",
                display_name="Vehicle Number",
                source="User Input",
            ))

        if not request.ewaybill_no or not request.ewaybill_no.strip():
            missing.append(MissingFieldResult(
                field_name="ewaybill_no",
                display_name="E-Waybill Number",
                source="User Input",
            ))

        # Check consignee data
        consignees = self.get_consignees()
        consignee = None
        for c in consignees:
            if c.get("Party Name", "").strip().lower() == request.party_name.strip().lower():
                consignee = c
                break

        if consignee is None:
            missing.append(MissingFieldResult(
                field_name="party_name",
                display_name="Party not found in Consignee Master",
                source="Consignee Master",
            ))
        else:
            consignee_fields = {
                "Address Line 1": "Dispatch Address Line 1",
                "Address Line 2": "Dispatch Address Line 2",
                "GSTIN": "GSTIN",
                "State": "State",
                "State Code": "State Code",
                "Dispatch From Address": "Dispatch From Address",
            }
            for field, display in consignee_fields.items():
                val = consignee.get(field, "")
                if not val or not str(val).strip():
                    missing.append(MissingFieldResult(
                        field_name=field,
                        display_name=display,
                        source="Consignee Master",
                    ))

        return missing

    def generate_missing_excel(
        self,
        request: DeliveryChallanRequest,
        missing_fields: List[MissingFieldResult],
    ) -> bytes:
        """Generate an Excel file listing missing fields for the user to fill in."""
        wb = Workbook()

        # --- Missing Fields Sheet ---
        ws_missing = wb.active
        ws_missing.title = "Missing Fields"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = ["Field Name", "Display Name", "Source", "Value (Fill This)"]
        for col, h in enumerate(headers, 1):
            cell = ws_missing.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, mf in enumerate(missing_fields, 2):
            ws_missing.cell(row=row_idx, column=1, value=mf.field_name).border = thin_border
            ws_missing.cell(row=row_idx, column=2, value=mf.display_name).border = thin_border
            ws_missing.cell(row=row_idx, column=3, value=mf.source).border = thin_border
            value_cell = ws_missing.cell(row=row_idx, column=4, value="")
            value_cell.border = thin_border
            value_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

        # Auto-fit column widths
        for col in ws_missing.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws_missing.column_dimensions[col[0].column_letter].width = max_length + 4

        # --- Challan Data Sheet ---
        ws_data = wb.create_sheet("Challan Data")
        data_headers = [
            "Challan Type", "Challan Date", "Party Name", "Vehicle No",
            "E-Waybill No", "HSN Code", "Rounded Off", "Remarks",
        ]
        for col, h in enumerate(data_headers, 1):
            cell = ws_data.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        data_values = [
            request.challan_type,
            str(request.challan_date),
            request.party_name,
            request.vehicle_no or "",
            request.ewaybill_no or "",
            request.hsn_code,
            request.rounded_off,
            request.remarks or "",
        ]
        for col, v in enumerate(data_values, 1):
            ws_data.cell(row=2, column=col, value=v).border = thin_border

        # --- Items Sheet ---
        ws_items = wb.create_sheet("Coil Items")
        item_headers = [
            "Coil No", "Description", "Grade", "Width", "Thk",
            "Coating", "Weight (kg)", "Weight (MT)", "Process",
            "HSN Code", "Rate per MT", "Value",
        ]
        for col, h in enumerate(item_headers, 1):
            cell = ws_items.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, item in enumerate(request.items, 2):
            values = [
                item.coil_number, item.description, item.grade,
                item.width, item.thk, item.coating,
                item.weight_kg, item.weight_mt, item.nature_of_processing,
                item.hsn_code, item.rate_per_mt, item.value,
            ]
            for col, v in enumerate(values, 1):
                ws_items.cell(row=row_idx, column=col, value=v).border = thin_border

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ──────────────────────────────────────────
    # Challan Number Generation
    # ──────────────────────────────────────────

    def generate_challan_number(self) -> str:
        df = self.google_service.get_worksheet_data(
            config.settings.api.google_sheets_id,
            config.DELIVERY_CHALLAN_SHEET,
        )

        numbers = []

        if "Challan Number" in df.columns:
            for val in df["Challan Number"].dropna():
                try:
                    num = int(val.split("/")[1])
                    numbers.append(num)
                except Exception:
                    continue

        next_num = max(numbers, default=0) + 1

        return f"F/{next_num:03}/{config.FINANCIAL_YEAR}"

    # ──────────────────────────────────────────
    # Create Challan (Save to Google Sheets)
    # ──────────────────────────────────────────

    def create_delivery_challan(
        self, request: DeliveryChallanRequest
    ) -> str:

        challan_number = self.generate_challan_number()

        total_weight_mt = sum(item.weight_mt for item in request.items)
        total_value = sum(item.value for item in request.items)

        record = DeliveryChallanRecord(
            challan_number=challan_number,
            challan_date=request.challan_date,
            challan_type=request.challan_type,
            party_name=request.party_name,
            vehicle_no=request.vehicle_no,
            ewaybill_no=request.ewaybill_no,
            hsn_code=request.hsn_code,
            total_weight_mt=total_weight_mt,
            total_value=total_value,
            rounded_off=request.rounded_off,
            remarks=request.remarks,
            coils_json=json.dumps(
                [item.dict() for item in request.items]
            ),
        )

        self.google_service.insert_row_before_last(
            config.settings.api.google_sheets_id,
            config.DELIVERY_CHALLAN_SHEET,
            [record.to_list()],
        )

        return challan_number

    # ──────────────────────────────────────────
    # PDF Generation (Existing — Preserved)
    # ──────────────────────────────────────────

    def generate_challan_pdf(self, challan_number: str) -> bytes:

        df = self.google_service.get_worksheet_data(
            config.settings.api.google_sheets_id,
            config.DELIVERY_CHALLAN_SHEET,
        )

        df.columns = df.columns.astype(str).str.strip()

        record = df[df["Challan Number"] == challan_number].iloc[0]
        items = json.loads(record["Coils JSON"])

        challan_date = datetime.strptime(
        str(record["Challan Date"]), "%Y-%m-%d"
        ).strftime("%d/%m/%Y")

        pdf = FPDF()
        pdf.set_auto_page_break(False)

        # 3 COPIES
        for _ in range(3):

            pdf.add_page()
            pdf.set_font("Arial", "", 10)

            # ===== HEADER =====
            pdf.cell(0, 6, f"To M/S.: {record['Party Name']}", ln=True)

            pdf.ln(2)

            pdf.set_font("Arial", "B", 10)
            pdf.cell(0, 6, "AMBA ENTERPRISE LTD. (UNIT I)", ln=True)

            pdf.set_font("Arial", "", 9)
            pdf.multi_cell(0, 5,
                "S. No 132 , H.No 1/4/1, Premraj Ind Est,\n"
                "Shed No B-1/2/3/4, Dalvi Wadi,\n"
                "Nanded Phata, Dhairy, Pune 411041"
            )

            pdf.ln(3)

            # ===== TITLE =====
            pdf.set_font("Arial", "B", 11)
            pdf.cell(0, 6, "Delivery Challan/ Receipt Inspection Report", ln=True)

            pdf.ln(3)

            # ===== DETAILS =====
            pdf.set_font("Arial", "", 9)
            pdf.cell(0, 5, f"Challan No: {challan_number}", ln=True)
            pdf.cell(0, 5, f"Challan Date: {challan_date}", ln=True)
            pdf.cell(0, 5, f"Vehicle No: {record['Vehicle No'] or ''}", ln=True)

            pdf.ln(4)

            # ===== TABLE =====
            pdf.set_font("Arial", "B", 9)

            headers = ["Job No", "P O No", "Particulars", "Gross Wt", "Deduction", "Net Wt"]
            widths = [30, 30, 40, 25, 25, 25]

            for h, w in zip(headers, widths):
                pdf.cell(w, 6, h, border=1)
            pdf.ln()

            pdf.set_font("Arial", "", 9)

            total_gross = 0
            total_net = 0

            for i, item in enumerate(items, start=1):

                gross = item.get("weight_mt", 0)
            net = gross

            total_gross += gross
            total_net += net

            pdf.cell(30, 6, str(10000 + i), border=1)
            pdf.cell(30, 6, f"{item.get('width', '')}", border=1)
            pdf.cell(40, 6, item.get("description", ""), border=1)
            pdf.cell(25, 6, f"{gross:.2f}", border=1)
            pdf.cell(25, 6, "-", border=1)
            pdf.cell(25, 6, f"{net:.2f}", border=1)
            pdf.ln()

            # ===== TOTAL =====
            pdf.set_font("Arial", "B", 9)
            pdf.cell(100, 6, "Total", border=1)
            pdf.cell(25, 6, f"{total_gross:.3f}", border=1)
            pdf.cell(25, 6, "0.000", border=1)
            pdf.cell(25, 6, f"{total_net:.3f}", border=1)
            pdf.ln()

            pdf.ln(5)

            # ===== FOOTER =====
            pdf.set_font("Arial", "", 9)

            pdf.cell(0, 5, "Note:", ln=True)
            pdf.cell(0, 5, "E. & O.E.", ln=True)

            pdf.multi_cell(0, 5,
            "Mentioned product(s) are received in the good Condition at our side."
            )

            pdf.ln(8)

            pdf.cell(90, 5, "Receiver's Signature", ln=0)
            pdf.cell(90, 5, "For AMBA ENTERPRISE LTD. (UNIT I)", ln=True)

            pdf.ln(8)

            pdf.cell(90, 5, "Prepared By", ln=0)
            pdf.cell(90, 5, "Verified By", ln=True)

        return bytes(pdf.output(dest="S"))